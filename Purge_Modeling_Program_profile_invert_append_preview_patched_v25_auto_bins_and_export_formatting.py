# Pipeline Pigging Simulation Program with Smart Enhancements (v8g_fix3)
import numpy as np
import pandas as pd
import zipfile
from lxml import etree as ET
from scipy.interpolate import CubicSpline
from geopy.distance import geodesic
import tkinter as tk
from tkinter import messagebox, filedialog, ttk, simpledialog
import matplotlib.pyplot as plt
try:
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
    from matplotlib.figure import Figure
except Exception:
    FigureCanvasTkAgg = None
    NavigationToolbar2Tk = None
    Figure = None
import time
import logging
import unittest
import os
import sys
import math
import re

# Set up logging with console output
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('purge_simulation.log'),
        logging.StreamHandler(sys.stdout)
    ]
)

# Test tkinter availability
try:
    tk.Tk().destroy()
    logging.info("Tkinter initialized successfully")
except Exception as e:
    logging.error(f"Tkinter initialization failed: {str(e)}")
    print(f"Error: Tkinter initialization failed: {str(e)}")
    sys.exit(1)

# NPS data (Outer Diameter in inches)
nps_data = {
    "1/8": {"OD_in": 0.405}, "1/4": {"OD_in": 0.540}, "3/8": {"OD_in": 0.675}, "1/2": {"OD_in": 0.840},
    "3/4": {"OD_in": 1.050}, "1": {"OD_in": 1.315}, "1 1/4": {"OD_in": 1.660}, "1 1/2": {"OD_in": 1.900},
    "2": {"OD_in": 2.375}, "2 1/2": {"OD_in": 2.875}, "3": {"OD_in": 3.500}, "3 1/2": {"OD_in": 4.000},
    "4": {"OD_in": 4.500}, "5": {"OD_in": 5.563}, "6": {"OD_in": 6.625}, "8": {"OD_in": 8.625},
    "10": {"OD_in": 10.750}, "12": {"OD_in": 12.750}, "14": {"OD_in": 14.000}, "16": {"OD_in": 16.000},
    "18": {"OD_in": 18.000}, "20": {"OD_in": 20.000}, "24": {"OD_in": 24.000}, "26": {"OD_in": 26.000},
    "28": {"OD_in": 28.000}, "30": {"OD_in": 30.000}, "32": {"OD_in": 32.000}, "34": {"OD_in": 34.000},
    "36": {"OD_in": 36.000}, "40": {"OD_in": 40.000}, "42": {"OD_in": 42.000}, "44": {"OD_in": 44.000},
    "48": {"OD_in": 48.000}, "52": {"OD_in": 52.000}, "56": {"OD_in": 56.000}, "60": {"OD_in": 60.000},
    "64": {"OD_in": 64.000}, "68": {"OD_in": 68.000}, "72": {"OD_in": 72.000}, "80": {"OD_in": 80.000},}

# Roughness data (in feet)
roughness_data = {
    1: {"material": "New Welded Steel", "roughness_ft": 0.00015},
    2: {"material": "Rusted/Corroded Welded Steel", "roughness_ft": 0.0005},
    3: {"material": "Welded HDPE", "roughness_ft": 0.000005}
}

# Fluid data
fluid_data = {
    1: {"name": "Diesel", "sg": 0.84, "viscosity_cst": 2.7},
    2: {"name": "Gasoline", "sg": 0.74, "viscosity_cst": 0.6},
    3: {"name": "Crude Oil", "sg": None, "viscosity_cst": None},
    4: {"name": "Water", "sg": 1.0, "viscosity_cst": 1.0},
    5: {"name": "NGL (Y1-grade)", "sg": 0.6, "viscosity_cst": 0.3}
}

# Nitrogen properties (used for gas-side calculations)
# We treat N2 as a real gas via a Peng–Robinson EOS for Z-factor.
N2_MW_KG_PER_MOL = 0.0280134
N2_CRIT_T_K = 126.192
N2_CRIT_P_PA = 3.3958e6  # ~492.3 psia
N2_ACENTRIC = 0.0372

# Assumed nitrogen temperature in the pipe (ground temperature), per user convention.
N2_TEMP_F_DEFAULT = 45.0

# Dynamic viscosity of nitrogen (Pa*s). A constant is adequate for purge-speed screening.
N2_VISCOSITY_PA_S = 1.75e-5
SCF_TO_FT3 = 1.0  # 1 SCF == 1 ft³ at STP
ATM_PSI = 14.7

# ------------------------ Small utility helpers ------------------------
def resolve_nps_key(nps_val):
    """Map GUI 'nps' (str like '16' or float 16.0) to a key present in nps_data."""
    # Exact match first
    if nps_val in nps_data:
        return nps_val
    # Try string->float->string conversions
    try:
        as_float = float(nps_val)
        # Try exact string of int
        cand = str(int(round(as_float)))
        if cand in nps_data:
            return cand
        # Try scanning keys numerically
        for k in nps_data.keys():
            try:
                if abs(float(k) - as_float) < 1e-9:
                    return k
            except Exception:
                continue
    except Exception:
        pass
    # As a last resort just string it
    k = str(nps_val)
    if k in nps_data:
        return k
    raise KeyError(f"NPS '{nps_val}' not found in nps_data; available keys: {list(nps_data.keys())}")

def print_inputs(inputs):
    print("\n=== Simulation Inputs ===")
    for key, value in sorted(inputs.items()):
        print(f"{key}: {value}")
    print("=== End Inputs ===\n")

def haversine_distance(lat1, lon1, lat2, lon2):
    R = 3958.8  # Earth radius in miles
    lat1, lon1, lat2, lon2 = map(np.radians, [lat1, lon1, lat2, lon2])
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = np.sin(dlat/2)**2 + np.cos(lat1) * np.cos(lat2) * np.sin(dlon/2)**2
    c = 2 * np.arcsin(np.sqrt(a))
    return R * c


# ------------------------ KML/KMZ coordinate extraction helpers ------------------------
def _lname(tag):
    """Return localname for an XML tag (strip namespace)."""
    try:
        if isinstance(tag, str) and '}' in tag:
            return tag.split('}', 1)[1]
    except Exception:
        pass
    return tag

def _get_secure_lxml_parser():
    """Create an lxml XMLParser with safer defaults (no external entities, no network)."""
    # Note: huge_tree=True helps with large KMLs; we still apply a point cap/decimation later.
    try:
        return ET.XMLParser(resolve_entities=False, no_network=True, recover=True, huge_tree=True)
    except TypeError:
        # Some lxml builds may not accept all flags
        return ET.XMLParser(resolve_entities=False, recover=True)

def _parse_kml_coordinates_text(coord_text):
    """Parse standard KML <coordinates> text tokens 'lon,lat,alt' separated by any whitespace."""
    out = []
    for token in (coord_text or '').strip().split():
        try:
            parts = token.split(',')
            if len(parts) < 2:
                continue
            lon = float(parts[0])
            lat = float(parts[1])
            elev = float(parts[2]) if len(parts) > 2 and parts[2] != '' else 0.0
            out.append((lat, lon, elev))
        except Exception:
            continue
    return out

def _parse_gx_coord_text(coord_text):
    """Parse Google Earth gx:coord text 'lon lat alt'."""
    try:
        parts = (coord_text or '').strip().split()
        if len(parts) < 2:
            return None
        lon = float(parts[0])
        lat = float(parts[1])
        elev = float(parts[2]) if len(parts) > 2 else 0.0
        return (lat, lon, elev)
    except Exception:
        return None

def extract_pipeline_coords_from_kml_root(root, file_path=None, multiline_mode='longest', max_points=200000):
    """
    Extract a best-effort pipeline centerline coordinate sequence from a KML root.

    Preference order:
      1) LineString <coordinates> (choose longest by default)
      2) gx:Track gx:coord (choose longest)
      3) Fallback: largest <coordinates> block not inside Polygon/LinearRing
    """
    # Collect LineString candidates
    linestring_candidates = []
    for ls in root.iter():
        if _lname(ls.tag) == 'LineString':
            # In lxml we can safely search within this subtree
            for ch in ls.iter():
                if _lname(ch.tag) == 'coordinates' and getattr(ch, 'text', None) and ch.text.strip():
                    pts = _parse_kml_coordinates_text(ch.text)
                    if pts:
                        linestring_candidates.append(pts)
                    break

    if linestring_candidates:
        if multiline_mode == 'all':
            coords = [p for seg in linestring_candidates for p in seg]
        else:
            coords = max(linestring_candidates, key=len)
        try:
            logging.info(f"KML parse: found {len(linestring_candidates)} LineString(s); using {len(coords)} points (mode={multiline_mode})")
        except Exception:
            pass
        # Point cap / decimation
        if max_points and len(coords) > max_points:
            step = max(1, int(len(coords) / max_points))
            coords = coords[::step]
            try:
                logging.warning(f"KML parse: decimated to {len(coords)} points (step={step}) for performance")
            except Exception:
                pass
        return coords

    # Collect gx:Track candidates (Google Earth)
    track_candidates = []
    for tr in root.iter():
        if _lname(tr.tag) == 'Track':
            pts = []
            for ch in tr.iter():
                if _lname(ch.tag) == 'coord' and getattr(ch, 'text', None) and ch.text.strip():
                    p = _parse_gx_coord_text(ch.text)
                    if p is not None:
                        pts.append(p)
            if pts:
                track_candidates.append(pts)

    if track_candidates:
        coords = max(track_candidates, key=len)
        try:
            logging.info(f"KML parse: found {len(track_candidates)} gx:Track(s); using {len(coords)} points")
        except Exception:
            pass
        if max_points and len(coords) > max_points:
            step = max(1, int(len(coords) / max_points))
            coords = coords[::step]
            try:
                logging.warning(f"KML parse: decimated to {len(coords)} points (step={step}) for performance")
            except Exception:
                pass
        return coords

    # Fallback: any <coordinates> blocks not inside Polygon/LinearRing
    fallback_candidates = []
    for ch in root.iter():
        if _lname(ch.tag) == 'coordinates' and getattr(ch, 'text', None) and ch.text.strip():
            # Skip polygon rings (to avoid pulling in area features)
            skip = False
            try:
                parent = ch.getparent()
                while parent is not None:
                    ln = _lname(parent.tag)
                    if ln in ('Polygon', 'LinearRing'):
                        skip = True
                        break
                    parent = parent.getparent()
            except Exception:
                # If parent traversal isn't available, just proceed (best-effort)
                pass
            if skip:
                continue
            pts = _parse_kml_coordinates_text(ch.text)
            if pts:
                fallback_candidates.append(pts)

    if fallback_candidates:
        coords = max(fallback_candidates, key=len)
        try:
            logging.warning(f"KML parse: no LineString/Track found; using largest <coordinates> block with {len(coords)} points")
        except Exception:
            pass
        if max_points and len(coords) > max_points:
            step = max(1, int(len(coords) / max_points))
            coords = coords[::step]
            try:
                logging.warning(f"KML parse: decimated to {len(coords)} points (step={step}) for performance")
            except Exception:
                pass
        return coords

    return []

def calculate_trendline_slope(mileposts, elevations, start_mp, end_mp):
    logging.debug(f"Calculating trendline slope from {start_mp} to {end_mp}")
    mask = (mileposts >= start_mp) & (mileposts <= end_mp)
    selected_mileposts = mileposts[mask]
    selected_elevations = elevations[mask]
    if len(selected_mileposts) < 2:
        logging.warning("Insufficient points for trendline slope calculation")
        return 0
    slope = (selected_elevations[-1] - selected_elevations[0]) / (selected_mileposts[-1] - selected_mileposts[0])
    logging.debug(f"Calculated slope: {slope}")
    return slope

def show_help():
    logging.info("Displaying help guide")
    help_text = """
Pipeline Pigging Simulation Program
==================================
- This version auto-adjusts N2 injection rate to maximize pig speed (<= Max) at a fixed exit pressure.
- Exit pressure tapers from 'run' to 'end' over the last N miles (Throttle Down).
- Friction uses Darcy–Weisbach (Swamee–Jain/Haaland), applied to remaining liquid slug length.
- Export fixes: NPS key is resolved robustly, avoiding the 16.0 KeyError.
    """
    try:
        help_window = tk.Toplevel()
        help_window.title("User Help")
        help_window.update()
        text = tk.Text(help_window, height=20, width=80)
        text.insert(tk.END, help_text)
        text.pack(padx=10, pady=10)
        tk.Button(help_window, text="Close", command=help_window.destroy).pack(pady=5)
        logging.info("Help guide displayed")
    except Exception as e:
        logging.error(f"Failed to display help guide: {str(e)}")
        print(f"Error: Failed to display help guide: {str(e)}")
        raise

def get_user_inputs(dialog_root, profile_start_mp, profile_end_mp):
    logging.info("Entering get_user_inputs")
    inputs = {}
    def validate_and_submit():
        try:
            nonlocal inputs
            logging.debug("Validating user inputs")
            nps = (nps_var.get() or "").strip()
            if not nps:
                raise ValueError("Nominal Pipe Size must be selected.")
            pipe_wt = wt_entry.get().strip()
            if not pipe_wt:
                raise ValueError("Wall thickness must be provided.")
            pipe_wt = float(pipe_wt)
            pipe_od = nps_data[resolve_nps_key(nps)]["OD_in"]
            if pipe_wt <= 0 or pipe_wt >= pipe_od / 2:
                raise ValueError("Wall thickness must be positive and less than half the OD.")
            roughness_num = int(material_var.get().split(':')[0])
            if roughness_num not in roughness_data:
                raise ValueError("Invalid pipe material selected.")
            fluid_num = int(fluid_var.get().split(':')[0])
            if fluid_num not in fluid_data:
                raise ValueError("Invalid fluid selected.")
            api_gravity = float(api_entry.get()) if fluid_num == 3 else None
            viscosity_cst = float(viscosity_entry.get()) if fluid_num == 3 else None
            if fluid_num == 3:
                if api_gravity <= 0:
                    raise ValueError("API Gravity must be positive.")
                if viscosity_cst <= 0:
                    raise ValueError("Viscosity must be positive.")
            # Optional max cap on N2 rate
            max_n2_rate_scfm = float(max_rate_entry.get()) if max_rate_entry.get().strip() else None
            n2_cutoff_scf = float(n2_cutoff_entry.get()) if n2_cutoff_entry.get().strip() else None
            exit_run = float(exit_run_entry.get())
            if exit_run < 0:
                raise ValueError("Minimum Exit Pressure (Run) must be non-negative.")
            exit_end = float(exit_end_entry.get())
            if exit_end < 0:
                raise ValueError("Exit Pressure (End) must be non-negative.")

            exit_behavior = exit_behavior_var.get().strip()
            if not exit_behavior:
                exit_behavior = "taper_last_n_miles"

            endpoint_constraint_mode = endpoint_constraint_mode_var.get().strip().lower()
            if not endpoint_constraint_mode:
                endpoint_constraint_mode = "none"
            if endpoint_constraint_mode not in ("none", "clamp_to_monitors"):
                endpoint_constraint_mode = "none"

            exit_pressure_min_clamp = float(exit_min_clamp_entry.get()) if exit_min_clamp_entry.get().strip() else None
            if exit_pressure_min_clamp is not None and exit_pressure_min_clamp < 0:
                raise ValueError("Exit Pressure Min Clamp must be non-negative if provided.")

            exit_pressure_max_clamp = float(exit_max_clamp_entry.get()) if exit_max_clamp_entry.get().strip() else None
            if exit_pressure_max_clamp is not None and exit_pressure_max_clamp < 0:
                raise ValueError("Exit Pressure Max Clamp must be non-negative if provided.")

            if (exit_pressure_min_clamp is not None) and (exit_pressure_max_clamp is not None) and (exit_pressure_min_clamp > exit_pressure_max_clamp):
                raise ValueError("Exit Pressure Min Clamp cannot exceed Exit Pressure Max Clamp.")

            exit_pressure_ramp_psi_per_hr = float(exit_ramp_entry.get()) if exit_ramp_entry.get().strip() else None
            if exit_pressure_ramp_psi_per_hr is not None and exit_pressure_ramp_psi_per_hr <= 0:
                raise ValueError("Exit Pressure Ramp Limit must be positive if provided.")

            max_outlet_flow_bph = float(max_outlet_flow_entry.get()) if max_outlet_flow_entry.get().strip() else None
            if max_outlet_flow_bph is not None and max_outlet_flow_bph <= 0:
                raise ValueError("Max Outlet Flow to Tankage must be positive if provided.")

            max_speed_from_flow_mph = None
            if max_outlet_flow_bph is not None:
                # Convert BPH to equivalent pig speed (mph) using the computed pipe ID
                pipe_id_in = pipe_od - 2 * pipe_wt
                pipe_diameter_ft = pipe_id_in / 12.0
                area_ft2 = math.pi * (pipe_diameter_ft / 2.0) ** 2
                q_ft3_s = max_outlet_flow_bph * 5.614583333333333 / 3600.0
                v_ft_s = q_ft3_s / area_ft2
                max_speed_from_flow_mph = v_ft_s * 3600.0 / 5280.0

            slack_pressure_psig = float(slack_entry.get()) if slack_entry.get().strip() else 50.0
            if slack_pressure_psig < 0:
                raise ValueError("Slack Threshold must be non-negative.")

            maop_psig = float(maop_entry.get()) if maop_entry.get().strip() else None
            if maop_psig is not None and maop_psig <= 0:
                raise ValueError("MAOP must be positive if provided.")

            n2_end = float(n2_end_entry.get())
            if n2_end < 0:
                raise ValueError("Estimated nitrogen pressure must be non-negative.")
            max_speed = float(max_speed_entry.get())
            if max_speed <= 0:
                raise ValueError("Max Pig Speed must be positive.")
            min_speed = float(min_speed_entry.get())
            if min_speed < 0:
                raise ValueError("Min Pig Speed must be non-negative.")
            if max_speed <= min_speed / 1.25:
                raise ValueError("Max Pig Speed must be greater than Min Pig Speed / 1.25.")
            target_speed = float(target_speed_entry.get())
            if target_speed <= 0 or target_speed > max_speed:
                raise ValueError("Target Pig Speed must be positive and <= Max Pig Speed.")
            purge_start = float(purge_start_entry.get())
            purge_end = float(purge_end_entry.get())
            system_end = float(system_end_entry.get())
            if not (purge_start < purge_end <= system_end):
                raise ValueError("Mileposts must satisfy: Start < End <= System End.")
            if purge_start < profile_start_mp:
                logging.warning(f"Purge Start Milepost {purge_start:.1f} is below profile range {profile_start_mp:.1f}. Adjusting to {profile_start_mp:.1f}.")
                purge_start = profile_start_mp
            if system_end > profile_end_mp + 1e-6:
                logging.warning(f"System Endpoint Milepost {system_end:.3f} exceeds profile range {profile_end_mp:.3f}. Adjusting to {profile_end_mp:.3f}.")
                system_end = profile_end_mp
            if purge_end > system_end:
                logging.warning(f"Purge End Milepost {purge_end:.1f} exceeds System Endpoint {system_end:.1f}. Adjusting to {system_end:.1f}.")
                purge_end = system_end
            throttle_down = float(throttle_down_entry.get())
            if throttle_down < 0 or throttle_down > (purge_end - purge_start):
                raise ValueError(f"Throttle Down Point must be between 0 and {purge_end - purge_start:.1f} miles.")
            
            has_ips = ips_var.get()
            ips_mp = None
            ips_mps = []
            ips_shutdown_dist = None
            min_pump_suction_pressure = None
            min_pump_flow_bph = None
            ips_check_dp_psi = 5.0
            if has_ips:
                ips_mp = ips_mp_entry.get().strip()
                if not ips_mp:
                    raise ValueError("IPS milepost(s) must be provided.")
                ips_mps = parse_milepost_list(ips_mp)
                if not ips_mps:
                    raise ValueError("Could not parse IPS milepost(s). Use comma-separated mileposts.")
                for _mp in ips_mps:
                    if not (purge_start < _mp < system_end):
                        raise ValueError("Each IPS milepost must be between Purge Start and System End.")
                ips_shutdown_dist = ips_shutdown_entry.get().strip()
                if not ips_shutdown_dist:
                    raise ValueError("IPS shutdown distance must be provided.")
                ips_shutdown_dist = float(ips_shutdown_dist)
                if ips_shutdown_dist <= 0:
                    raise ValueError("IPS shutdown distance must be positive.")
                min_pump_suction_pressure = min_pump_suction_entry.get().strip()
                if not min_pump_suction_pressure:
                    raise ValueError("Minimum Pump Suction Pressure must be provided.")
                min_pump_suction_pressure = float(min_pump_suction_pressure)
                if min_pump_suction_pressure < 0:
                    raise ValueError("Minimum Pump Suction Pressure must be non-negative.")

                # Optional: station minimum mainline flow (BPH). If provided and positive, station is only considered "in use"
                # when mainline flow meets/exceeds this value (protects against unrealistic low-flow operation without modeling recirc).
                min_pump_flow_bph = None
                _mpf = min_pump_flow_entry.get().strip()
                if _mpf:
                    min_pump_flow_bph = float(_mpf)
                    if min_pump_flow_bph <= 0:
                        raise ValueError("Minimum Pump Flow must be positive if provided.")

                # Optional: check valve seal threshold. If pump ΔP is below this, assume the check does not fully seal and the station
                # behaves like flow-through (not a hydraulic boundary).
                _cdp = ips_check_dp_entry.get().strip()
                ips_check_dp_psi = float(_cdp) if _cdp else 5.0
                if ips_check_dp_psi < 0:
                    raise ValueError("IPS check valve seal ΔP must be non-negative.")
            
            resolution = int(resolution_entry.get()) if resolution_entry.get().strip() else 500
            if resolution < 50 or resolution > 2000:
                raise ValueError("Resolution must be between 50 and 2000 points.")
            
            resample_profile_to_resolution = bool(resample_profile_var.get())

            # "Max Nitrogen Pressure" is the maximum allowable nitrogen pressure at the injection gauge.
            # For backwards compatibility with older files/exports, we also store it under
            # the legacy key name "max_drive_pressure".
            max_nitrogen_pressure = float(max_nitrogen_pressure_entry.get())

            inputs.update({
                'nps': nps, 'pipe_wt': pipe_wt, 'roughness_num': roughness_num, 'fluid_num': fluid_num,
                'api_gravity': api_gravity, 'viscosity_cst': viscosity_cst, 'max_n2_rate_scfm': max_n2_rate_scfm,
                'exit_pressure_run': exit_run, 'exit_pressure_end': exit_end, 'exit_pressure_behavior': exit_behavior, 'endpoint_constraint_mode': endpoint_constraint_mode, 'exit_pressure_min_clamp': exit_pressure_min_clamp, 'exit_pressure_max_clamp': exit_pressure_max_clamp, 'exit_pressure_ramp_psi_per_hr': exit_pressure_ramp_psi_per_hr, 'slack_pressure_psig': slack_pressure_psig, 'maop_psig': maop_psig, 'n2_end_pressure': n2_end,
                'max_pig_speed': max_speed, 'min_pig_speed': min_speed, 'target_pig_speed': target_speed,
                'purge_start_mp': purge_start, 'purge_end_mp': purge_end, 'system_end_mp': system_end,
                'throttle_down_miles': throttle_down,
                'elevation_format': elev_format_var.get(),
                'elevation_units': elev_units_var.get(), 'has_ips': has_ips, 'ips_mp': ips_mp, 'ips_mps': ips_mps,
                'ips_shutdown_dist': ips_shutdown_dist, 'min_pump_suction_pressure': min_pump_suction_pressure, 'min_pump_flow_bph': min_pump_flow_bph, 'ips_check_dp_psi': ips_check_dp_psi,
                                'max_outlet_flow_bph': max_outlet_flow_bph,
                'max_pig_speed_from_max_outlet_flow_mph': max_speed_from_flow_mph,
'resolution': resolution,
                'resample_profile_to_resolution': resample_profile_to_resolution,
                'n2_cutoff_scf': n2_cutoff_scf,
                'max_nitrogen_pressure': max_nitrogen_pressure,
                'max_drive_pressure': max_nitrogen_pressure
            })
            inputs.setdefault('temperature_f', float(N2_TEMP_F_DEFAULT))
            logging.info(f"Validated inputs: NPS={nps}, Max N2 Rate={(max_n2_rate_scfm if max_n2_rate_scfm is not None else 'auto')} SCFM, Target Speed={target_speed} mph, Purge Start={purge_start}, Purge End={purge_end}, Resolution={resolution}, ResampleProfile={resample_profile_to_resolution}")
            win.destroy()
        except ValueError as e:
            logging.error(f"Input validation failed: {str(e)}")
            messagebox.showerror("Input Error", str(e), parent=win)
    
    try:
        logging.info("Initializing Tkinter root for input GUI")
        win = tk.Toplevel(dialog_root)
        logging.info("Tkinter input GUI created")
        win.title("Pipeline Pigging Simulation Inputs")
        # Make the inputs window usable on smaller screens
        try:
            win.geometry("600x800")
        except Exception:
            pass
        win.minsize(520, 600)

        # Bring to front
        win.lift()
        win.attributes('-topmost', True)
        win.after(250, lambda: win.attributes('-topmost', False))
        win.update_idletasks()

        # --- Scrollable form (so Submit is always reachable) ---
        container = ttk.Frame(win)
        container.grid(row=0, column=0, sticky="nsew")
        win.rowconfigure(0, weight=1)
        win.columnconfigure(0, weight=1)

        canvas = tk.Canvas(container, highlightthickness=0, borderwidth=0)
        vbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vbar.set)

        canvas.grid(row=0, column=0, sticky="nsew")
        vbar.grid(row=0, column=1, sticky="ns")
        container.rowconfigure(0, weight=1)
        container.columnconfigure(0, weight=1)

        # 'root' becomes the inner frame that holds all widgets
        root = ttk.Frame(canvas)
        _form_window = canvas.create_window((0, 0), window=root, anchor="nw")

        def _sync_scrollregion(_event=None):
            canvas.configure(scrollregion=canvas.bbox("all"))

        def _sync_width(event):
            # Keep the inner frame width matched to the canvas width
            canvas.itemconfigure(_form_window, width=event.width)

        root.bind("<Configure>", _sync_scrollregion)
        canvas.bind("<Configure>", _sync_width)

        # Mouse wheel scrolling (Windows/macOS) + Linux
        def _on_mousewheel(event):
            try:
                canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            except Exception:
                pass

        def _on_mousewheel_linux(event):
            try:
                if event.num == 4:
                    canvas.yview_scroll(-3, "units")
                elif event.num == 5:
                    canvas.yview_scroll(3, "units")
            except Exception:
                pass

        win.bind("<MouseWheel>", _on_mousewheel)
        win.bind("<Button-4>", _on_mousewheel_linux)
        win.bind("<Button-5>", _on_mousewheel_linux)

        messagebox.showinfo(
            "Tip",
            "This input window is scrollable. Use the mouse wheel or the scrollbar on the right to reach Submit.",
            parent=win
        )
    except Exception as e:
        logging.error(f"Failed to initialize Tkinter input GUI: {str(e)}")
        print(f"Error: Failed to initialize Tkinter input GUI: {str(e)}")
        raise
    
    row = 0
    tk.Label(root, text="Nominal Pipe Size (NPS):").grid(row=row, column=0, sticky='e')
    nps_var = tk.StringVar(value="")
    nps_values = [""] + list(nps_data.keys())
    nps_combo = ttk.Combobox(root, textvariable=nps_var, values=nps_values, state="readonly")
    nps_combo.grid(row=row, column=1, sticky='w')
    row += 1
    
    tk.Label(root, text="Wall Thickness (in):").grid(row=row, column=0, sticky='e')
    wt_entry = tk.Entry(root)
    wt_entry.insert(0, "0.280")
    wt_entry.grid(row=row, column=1)
    row += 1
    
    tk.Label(root, text="Pipe Material:").grid(row=row, column=0, sticky='e')
    material_var = tk.StringVar(value="2: Rusted/Corroded Welded Steel")
    material_options = [f"{k}: {v['material']}" for k, v in roughness_data.items()]
    ttk.Combobox(root, textvariable=material_var, values=material_options).grid(row=row, column=1)
    row += 1
    
    tk.Label(root, text="Fluid Type:").grid(row=row, column=0, sticky='e')
    fluid_var = tk.StringVar(value="3: Crude Oil")
    fluid_options = [f"{k}: {v['name']}" for k, v in fluid_data.items()]
    ttk.Combobox(root, textvariable=fluid_var, values=fluid_options).grid(row=row, column=1)
    row += 1
    
    tk.Label(root, text="API Gravity (if Crude Oil):").grid(row=row, column=0, sticky='e')
    api_entry = tk.Entry(root)
    api_entry.insert(0, "25")
    api_entry.grid(row=row, column=1)
    row += 1
    
    tk.Label(root, text="Viscosity (cSt) (if Crude Oil):").grid(row=row, column=0, sticky='e')
    viscosity_entry = tk.Entry(root)
    viscosity_entry.insert(0, "200")
    viscosity_entry.grid(row=row, column=1)
    row += 1
    
    tk.Label(root, text="Min Exit Pressure (Run) (psi):").grid(row=row, column=0, sticky='e')
    exit_run_entry = tk.Entry(root)
    exit_run_entry.insert(0, "100")
    exit_run_entry.grid(row=row, column=1)
    row += 1
    
    tk.Label(root, text="Exit Pressure (End) (psi):").grid(row=row, column=0, sticky='e')
    exit_end_entry = tk.Entry(root)
    exit_end_entry.insert(0, "100")
    exit_end_entry.grid(row=row, column=1)
    row += 1

    # Exit pressure boundary condition behavior (how the downstream vent/backpressure is operated)
    tk.Label(root, text="Exit Pressure Behavior:").grid(row=row, column=0, sticky='e')
    exit_behavior_var = tk.StringVar(value="taper_last_n_miles")
    exit_behavior_combo = ttk.Combobox(
        root,
        textvariable=exit_behavior_var,
        state="readonly",
        values=[
            "taper_last_n_miles",
            "step_last_n_miles",
            "linear_ramp",
            "constant_run",
            "constant_end",
        ]
    )
    exit_behavior_combo.grid(row=row, column=1, sticky='w')
    row += 1

    # Endpoint constraint scheme (optional). This is a *clamp* on the downstream BC.
    tk.Label(root, text="Endpoint Constraint Mode:").grid(row=row, column=0, sticky='e')
    endpoint_constraint_mode_var = tk.StringVar(value="none")
    endpoint_constraint_combo = ttk.Combobox(
        root,
        textvariable=endpoint_constraint_mode_var,
        state="readonly",
        values=[
            "none",
            "clamp_to_monitors",
        ]
    )
    endpoint_constraint_combo.grid(row=row, column=1, sticky='w')
    row += 1

    tk.Label(root, text="Exit Pressure Min Clamp (psig) [optional]:").grid(row=row, column=0, sticky='e')
    exit_min_clamp_entry = tk.Entry(root)
    exit_min_clamp_entry.insert(0, "")
    exit_min_clamp_entry.grid(row=row, column=1)
    row += 1

    tk.Label(root, text="Exit Pressure Max Clamp (psig) [optional]:").grid(row=row, column=0, sticky='e')
    exit_max_clamp_entry = tk.Entry(root)
    exit_max_clamp_entry.insert(0, "")
    exit_max_clamp_entry.grid(row=row, column=1)
    row += 1

    tk.Label(root, text="Exit Pressure Ramp Limit (psi/hr) [optional]:").grid(row=row, column=0, sticky='e')
    exit_ramp_entry = tk.Entry(root)
    exit_ramp_entry.insert(0, "")
    exit_ramp_entry.grid(row=row, column=1)
    row += 1
    tk.Label(root, text="Max Outlet Flow to Tankage (BPH) [optional]:").grid(row=row, column=0, sticky='e')
    max_outlet_flow_entry = tk.Entry(root)
    max_outlet_flow_entry.insert(0, "")
    max_outlet_flow_entry.grid(row=row, column=1)
    row += 1


    # Monitoring thresholds
    tk.Label(root, text="Slack Threshold (psig):").grid(row=row, column=0, sticky='e')
    slack_entry = tk.Entry(root)
    slack_entry.insert(0, "50")
    slack_entry.grid(row=row, column=1)
    row += 1

    tk.Label(root, text="MAOP (psig) (optional):").grid(row=row, column=0, sticky='e')
    maop_entry = tk.Entry(root)
    maop_entry.insert(0, "")
    maop_entry.grid(row=row, column=1)
    row += 1

    tk.Label(root, text="Est. N2 Pressure at End (psi):").grid(row=row, column=0, sticky='e')
    n2_end_entry = tk.Entry(root)
    n2_end_entry.insert(0, "300")
    n2_end_entry.grid(row=row, column=1)
    row += 1

    # This is the maximum nitrogen pressure allowed at the injection gauge.
    tk.Label(root, text="Max Nitrogen Pressure (psi):").grid(row=row, column=0, sticky='e')
    max_nitrogen_pressure_entry = tk.Entry(root)
    max_nitrogen_pressure_entry.insert(0, "400")
    max_nitrogen_pressure_entry.grid(row=row, column=1)
    row += 1

    tk.Label(root, text="Max N2 Rate (SCFM) [optional]:").grid(row=row, column=0, sticky='e')
    max_rate_entry = tk.Entry(root)
    max_rate_entry.insert(0, "")
    max_rate_entry.grid(row=row, column=1)
    row += 1

    tk.Label(root, text="N2 Cutoff Volume (SCF) [optional]:").grid(row=row, column=0, sticky='e')
    n2_cutoff_entry = tk.Entry(root)
    n2_cutoff_entry.insert(0, "")
    n2_cutoff_entry.grid(row=row, column=1)
    row += 1
    
    tk.Label(root, text="Max Pig Speed (mph):").grid(row=row, column=0, sticky='e')
    max_speed_entry = tk.Entry(root)
    max_speed_entry.insert(0, "4")
    max_speed_entry.grid(row=row, column=1)
    row += 1
    
    tk.Label(root, text="Min Pig Speed (mph):").grid(row=row, column=0, sticky='e')
    min_speed_entry = tk.Entry(root)
    min_speed_entry.insert(0, "1")
    min_speed_entry.grid(row=row, column=1)
    row += 1
    
    tk.Label(root, text="Target Pig Speed (mph):").grid(row=row, column=0, sticky='e')
    target_speed_entry = tk.Entry(root)
    target_speed_entry.insert(0, "2")
    target_speed_entry.grid(row=row, column=1)
    row += 1
    
    tk.Label(root, text=f"Purge Start Milepost (min {profile_start_mp:.1f}):").grid(row=row, column=0, sticky='e')
    purge_start_entry = tk.Entry(root)
    purge_start_entry.insert(0, f"{profile_start_mp:.1f}")
    purge_start_entry.grid(row=row, column=1)
    row += 1
    
    tk.Label(root, text=f"Purge End Milepost (max {profile_end_mp:.1f}):").grid(row=row, column=0, sticky='e')
    purge_end_entry = tk.Entry(root)
    purge_end_entry.insert(0, f"{profile_end_mp:.1f}")
    purge_end_entry.grid(row=row, column=1)
    row += 1
    
    tk.Label(root, text=f"System Endpoint Milepost (max {profile_end_mp:.1f}):").grid(row=row, column=0, sticky='e')
    system_end_entry = tk.Entry(root)
    system_end_entry.insert(0, f"{profile_end_mp:.1f}")
    system_end_entry.grid(row=row, column=1)
    row += 1
    
    tk.Label(root, text="Throttle Down Point (miles from end):").grid(row=row, column=0, sticky='e')
    throttle_down_entry = tk.Entry(root)
    throttle_down_entry.insert(0, "9.6")
    throttle_down_entry.grid(row=row, column=1)
    row += 1
    
    tk.Label(root, text="Is there an Intermediate Pump Station?").grid(row=row, column=0, sticky='e')
    ips_var = tk.BooleanVar(value=False)
    tk.Checkbutton(root, variable=ips_var).grid(row=row, column=1)
    row += 1
    
    tk.Label(root, text="IPS Milepost(s) (comma-separated):").grid(row=row, column=0, sticky='e')
    ips_mp_entry = tk.Entry(root)
    ips_mp_entry.grid(row=row, column=1)
    row += 1
    
    tk.Label(root, text="IPS Shutdown Distance (miles):").grid(row=row, column=0, sticky='e')
    ips_shutdown_entry = tk.Entry(root)
    ips_shutdown_entry.insert(0, "1")
    ips_shutdown_entry.grid(row=row, column=1)
    row += 1
    
    tk.Label(root, text="Minimum Pump Suction Pressure (psi):").grid(row=row, column=0, sticky='e')
    min_pump_suction_entry = tk.Entry(root)
    min_pump_suction_entry.insert(0, "50")
    min_pump_suction_entry.grid(row=row, column=1)
    row += 1


    tk.Label(root, text="Minimum Pump Flow (BPH) [optional]:").grid(row=row, column=0, sticky='e')
    min_pump_flow_entry = tk.Entry(root)
    min_pump_flow_entry.grid(row=row, column=1)
    row += 1

    tk.Label(root, text="IPS Check Valve Seal ΔP (psi) [optional]:").grid(row=row, column=0, sticky='e')
    ips_check_dp_entry = tk.Entry(root)
    ips_check_dp_entry.insert(0, "5")
    ips_check_dp_entry.grid(row=row, column=1)
    row += 1
    
    tk.Label(root, text="Elevation Profile Format:").grid(row=row, column=0, sticky='e')
    elev_format_var = tk.StringVar(value="TXT")
    ttk.Combobox(root, textvariable=elev_format_var, values=["KMZ/KML", "Excel", "TXT"]).grid(row=row, column=1)
    row += 1
    
    tk.Label(root, text="Elevation Units:").grid(row=row, column=0, sticky='e')
    elev_units_var = tk.StringVar(value="Feet")
    ttk.Combobox(root, textvariable=elev_units_var, values=["Feet", "Meters"]).grid(row=row, column=1)
    row += 1
    
    tk.Label(root, text="Resolution (points):").grid(row=row, column=0, sticky='e')
    resolution_entry = tk.Entry(root)
    resolution_entry.insert(0, "500")
    resolution_entry.grid(row=row, column=1)
    row += 1
    
    # Purge profile resampling behavior
    # If enabled (recommended), the purge segment is resampled to exactly `resolution` points using a spline.
    # If disabled, the simulator uses the native uploaded points (within Purge Start/End), with start/end enforced.
    resample_profile_var = tk.BooleanVar(value=True)
    tk.Checkbutton(
        root,
        text="Resample purge segment to Resolution (recommended)",
        variable=resample_profile_var
    ).grid(row=row, column=0, columnspan=2, sticky='w')
    row += 1

    tk.Button(root, text="Submit", command=validate_and_submit).grid(row=row, column=0, columnspan=2)
    tk.Button(root, text="Help", command=show_help).grid(row=row, column=1, sticky='e')
    
    logging.info("Starting Tkinter wait for input GUI")
    win.grab_set()
    # Block until the user closes the window via Submit/Cancel.
    # Tkinter's wait_window() keeps the UI responsive without arbitrary timeouts.
    win.wait_window()

    logging.info("Exiting get_user_inputs")
    return inputs

def calculate_friction_loss(v, L, D, fluid_density, viscosity, roughness, purged_fraction=0):
    # Convert "fluid_density" (weight density, lb/ft^3) to mass density (slug/ft^3)
    rho_m = fluid_density / 32.174
    # NOTE: Downstream of a perfectly-sealing pig we assume the segment is liquid-full.
    # Therefore, do NOT blend viscosity/density toward nitrogen based on purged fraction.
    mu_eff = viscosity
    try:
        Re = rho_m * v * D / max(mu_eff, 1e-12)
        # Haaland friction factor
        haaland = -1.8 * np.log10((roughness / D / 3.7)**1.11 + 6.9 / max(Re, 1e-12))
        f = (1.0 / haaland)**2
        # Darcy–Weisbach pressure drop (psi)
        dp_psf = f * (L / D) * 0.5 * rho_m * v * v
        friction_loss = dp_psf / 144.0
        return friction_loss
    except Exception as e:
        logging.error(f"Friction loss calculation failed: {str(e)}")
        raise ValueError(f"Error in friction loss calculation: {str(e)}")

def parse_milepost_list(val):
    """Parse a milepost list from a GUI string (comma/semicolon/space separated) or pass-through list/tuple."""
    if val is None:
        return []
    if isinstance(val, (list, tuple, np.ndarray)):
        out = []
        for x in val:
            try:
                out.append(float(x))
            except Exception:
                continue
        return out
    s = str(val).strip()
    if not s:
        return []
    # allow commas/semicolons/spaces
    parts = re.split(r"[;,\s]+", s)
    out = []
    for p in parts:
        if not p:
            continue
        try:
            out.append(float(p))
        except Exception:
            continue
    return out


def target_exit_pressure(mp, inputs, purge_start, purge_end, throttle_down, t_hours=None):
    """Return outlet/endpoint pressure (psig) based on a selectable behavior program.

    This is a boundary condition (BC) — not an active controller. It represents how the downstream
    operator/vent/backpressure regulator is being run.
    """
    run_p = float(inputs.get('exit_pressure_run', 0.0))
    end_p = float(inputs.get('exit_pressure_end', run_p))
    behavior = str(inputs.get('exit_pressure_behavior', 'taper_last_n_miles')).strip().lower()

    # normalize common aliases
    aliases = {
        'taper': 'taper_last_n_miles',
        'taper_last': 'taper_last_n_miles',
        'linear': 'linear_ramp',
        'ramp': 'linear_ramp',
        'hold_run': 'constant_run',
        'hold_end': 'constant_end',
        'step': 'step_last_n_miles',
    }
    behavior = aliases.get(behavior, behavior)

    # guard against divide-by-zero
    purge_len = max(1e-12, float(purge_end) - float(purge_start))
    dist_from_end = float(purge_end) - float(mp)

    if behavior == 'constant_end':
        return end_p
    if behavior == 'constant_run':
        return run_p

    if behavior == 'linear_ramp':
        # run at purge start -> end at purge end (independent of throttle_down)
        frac = (float(mp) - float(purge_start)) / purge_len
        frac = max(0.0, min(1.0, frac))
        return run_p * (1.0 - frac) + end_p * frac

    if behavior == 'step_last_n_miles':
        # hold run until within throttle_down miles of end, then step to end_p
        td = max(0.0, float(throttle_down or 0.0))
        if td > 0.0 and dist_from_end <= td:
            return end_p
        return run_p

    # default: taper over the last throttle_down miles
    td = max(0.0, float(throttle_down or 0.0))
    if td > 0.0 and dist_from_end <= td:
        taper_factor = max(0.0, min(1.0, dist_from_end / td))
        return run_p * taper_factor + end_p * (1.0 - taper_factor)
    return run_p


def _local_extrema_indices(y: np.ndarray):
    """Return (max_idx_list, min_idx_list) for local extrema in a 1D array."""
    y = np.asarray(y, dtype=float)
    if y.size < 3:
        return [], []
    dy = np.diff(y)
    # Replace zeros with nearest non-zero sign to avoid flat segments creating noise
    s = np.sign(dy)
    if np.all(s == 0):
        return [], []
    # forward fill then back fill zeros
    for i in range(1, s.size):
        if s[i] == 0:
            s[i] = s[i-1]
    for i in range(s.size - 2, -1, -1):
        if s[i] == 0:
            s[i] = s[i+1]
    max_idx = []
    min_idx = []
    for i in range(1, s.size):
        if s[i-1] > 0 and s[i] < 0:
            max_idx.append(i)
        elif s[i-1] < 0 and s[i] > 0:
            min_idx.append(i)
    return max_idx, min_idx


def _pick_with_spacing(cand_idx, mps, elevs, n_pick, min_spacing_mi, highest=True):
    """Pick up to n_pick candidate indices with minimum spacing in mileposts."""
    if not cand_idx:
        return []
    mps = np.asarray(mps, dtype=float)
    elevs = np.asarray(elevs, dtype=float)
    # sort candidates by elevation
    cand_idx = list(dict.fromkeys([int(i) for i in cand_idx if 0 <= int(i) < mps.size]))
    cand_idx.sort(key=lambda i: elevs[i], reverse=highest)
    chosen = []
    for i in cand_idx:
        mp = mps[i]
        if all(abs(mp - mps[j]) >= min_spacing_mi for j in chosen):
            chosen.append(i)
        if len(chosen) >= n_pick:
            break
    return chosen


def _binned_extrema_indices(mps: np.ndarray, elevs: np.ndarray, n_bins: int = 10):
    """Return (hi_idx_list, lo_idx_list) of per-bin extrema along milepost.

    Splits the profile range [mps[0], mps[-1]] into `n_bins` bins and selects the
    max-elevation and min-elevation index within each bin (if present).
    """
    mps = np.asarray(mps, dtype=float)
    elevs = np.asarray(elevs, dtype=float)
    if mps.size == 0:
        return [], []
    n_bins = int(max(1, n_bins))
    mp0 = float(mps[0])
    mp1 = float(mps[-1])
    if not np.isfinite(mp0) or not np.isfinite(mp1) or mp1 <= mp0:
        return [], []
    edges = np.linspace(mp0, mp1, n_bins + 1)
    hi_idx = []
    lo_idx = []
    for b in range(n_bins):
        lo = edges[b]
        hi = edges[b + 1]
        if b == n_bins - 1:
            mask = (mps >= lo) & (mps <= hi)
        else:
            mask = (mps >= lo) & (mps < hi)
        idx = np.where(mask)[0]
        if idx.size == 0:
            continue
        # Pick extrema within the bin
        e = elevs[idx]
        hi_i = int(idx[int(np.nanargmax(e))]) if np.any(np.isfinite(e)) else int(idx[0])
        lo_i = int(idx[int(np.nanargmin(e))]) if np.any(np.isfinite(e)) else int(idx[0])
        hi_idx.append(hi_i)
        lo_idx.append(lo_i)
    # De-dupe while preserving order
    hi_idx = list(dict.fromkeys(hi_idx))
    lo_idx = list(dict.fromkeys(lo_idx))
    return hi_idx, lo_idx


def _collapse_nearby(idx_list, mps, elevs, min_spacing_mi, highest=True, protect_idx=None):
    """Collapse picks closer than min_spacing_mi, keeping the more extreme point.

    idx_list is assumed to be indices into mps/elevs. Returns indices sorted by milepost.
    protect_idx (set) are indices that will always be kept when collapsing.
    """
    if not idx_list:
        return []
    protect_idx = set(protect_idx or [])
    mps = np.asarray(mps, dtype=float)
    elevs = np.asarray(elevs, dtype=float)
    idx = sorted(set(int(i) for i in idx_list if 0 <= int(i) < mps.size), key=lambda i: float(mps[i]))
    if min_spacing_mi is None:
        return idx
    try:
        min_spacing_mi = float(min_spacing_mi)
    except Exception:
        return idx
    if min_spacing_mi <= 0:
        return idx

    out = []
    for i in idx:
        if not out:
            out.append(i)
            continue
        prev = out[-1]
        if abs(float(mps[i]) - float(mps[prev])) >= min_spacing_mi:
            out.append(i)
            continue

        # Too close: decide which to keep
        if prev in protect_idx and i in protect_idx:
            # Keep both protected; allow adjacency
            out.append(i)
            continue
        if prev in protect_idx:
            continue
        if i in protect_idx:
            out[-1] = i
            continue

        # Neither protected: keep the more extreme by elevation
        if highest:
            keep_i = i if float(elevs[i]) >= float(elevs[prev]) else prev
        else:
            keep_i = i if float(elevs[i]) <= float(elevs[prev]) else prev
        out[-1] = keep_i

    # final unique preserve order
    out2 = []
    seen = set()
    for i in out:
        if i not in seen:
            out2.append(i); seen.add(i)
    return out2


def build_auto_monitor_points(mps, elevs, n_high=10, n_low=10, min_spacing_mi=0.5, name_prefix="", n_bins=10, mode="binned"):
    """Build monitor points at representative elevation highs/lows along a profile.

    Two modes:
      - mode='binned' (default): split profile into `n_bins` sections and select the
        max and min elevation within each section. Always includes the global max and min.
      - mode='local_extrema': find local extrema and then pick top N with spacing.

    Parameters
    ----------
    mps, elevs : array-like
        Milepost (miles) and elevation (ft) arrays for the profile.
    n_high, n_low : int
        In 'binned' mode: maximum number of highs/lows to keep after collapsing nearby points.
        In 'local_extrema' mode: number of local highs/lows to pick.
    min_spacing_mi : float
        Minimum spacing between picked points (miles). In binned mode, used to collapse
        near-duplicate picks at bin boundaries while preserving global extrema.
    name_prefix : str
        Optional prefix to prepend to point names/categories (e.g., 'SYS ' or 'PURGE ').
    n_bins : int
        Number of bins/sections when mode='binned'.
    mode : str
        'binned' or 'local_extrema'
    """
    mps = np.asarray(mps, dtype=float)
    elevs = np.asarray(elevs, dtype=float)
    points = []
    if mps.size < 2:
        return points

    scope = (name_prefix or '').strip()

    def _nm(kind, label, mp):
        if name_prefix:
            return f"{name_prefix}{kind} {label} @ MP {mp:.3f}"
        return f"{kind} {label} @ MP {mp:.3f}"

    mode = (mode or "binned").strip().lower()

    # Global extrema (always include)
    try:
        g_hi = int(np.nanargmax(elevs))
    except Exception:
        g_hi = int(np.argmax(elevs))
    try:
        g_lo = int(np.nanargmin(elevs))
    except Exception:
        g_lo = int(np.argmin(elevs))

    if mode in ("local", "local_extrema", "peaks"):
        max_idx, min_idx = _local_extrema_indices(elevs)
        max_idx = sorted(max_idx, key=lambda i: elevs[i], reverse=True)
        min_idx = sorted(min_idx, key=lambda i: elevs[i])
        high_sel = _pick_with_spacing(max_idx, mps, elevs, int(n_high), float(min_spacing_mi), highest=True)
        low_sel  = _pick_with_spacing(min_idx, mps, elevs, int(n_low),  float(min_spacing_mi), highest=False)
        # Ensure globals included
        if g_hi not in high_sel:
            high_sel = [g_hi] + high_sel
        if g_lo not in low_sel:
            low_sel = [g_lo] + low_sel
        # Collapse nearby w/ protection for global indices
        high_sel = _collapse_nearby(high_sel, mps, elevs, min_spacing_mi, highest=True, protect_idx={g_hi})
        low_sel  = _collapse_nearby(low_sel,  mps, elevs, min_spacing_mi, highest=False, protect_idx={g_lo})
        # Cap counts (keeping globals)
        if len(high_sel) > int(n_high):
            # keep global then next most extreme
            rest = [i for i in high_sel if i != g_hi]
            rest.sort(key=lambda i: float(elevs[i]), reverse=True)
            high_sel = [g_hi] + rest[:max(0, int(n_high)-1)]
        if len(low_sel) > int(n_low):
            rest = [i for i in low_sel if i != g_lo]
            rest.sort(key=lambda i: float(elevs[i]))
            low_sel = [g_lo] + rest[:max(0, int(n_low)-1)]
    else:
        # Binned mode: per-section extrema (captures “big hill then long descent” profiles better)
        hi_bins, lo_bins = _binned_extrema_indices(mps, elevs, n_bins=int(n_bins))
        high_sel = list(dict.fromkeys([g_hi] + hi_bins))
        low_sel  = list(dict.fromkeys([g_lo] + lo_bins))

        # Collapse near-duplicates at bin boundaries while protecting global extrema
        high_sel = _collapse_nearby(high_sel, mps, elevs, min_spacing_mi, highest=True, protect_idx={g_hi})
        low_sel  = _collapse_nearby(low_sel,  mps, elevs, min_spacing_mi, highest=False, protect_idx={g_lo})

        # Optional caps (keep representation; if capped, keep one-per-bin feel by milepost order)
        try:
            n_high = int(n_high)
        except Exception:
            n_high = 10
        try:
            n_low = int(n_low)
        except Exception:
            n_low = 10

        if n_high > 0 and len(high_sel) > n_high:
            # keep global high, then spread remaining by milepost (every k-th)
            rest = [i for i in high_sel if i != g_hi]
            if n_high == 1:
                high_sel = [g_hi]
            else:
                k = max(1, int(math.ceil(len(rest) / float(n_high - 1))))
                high_sel = [g_hi] + rest[::k][:n_high-1]

        if n_low > 0 and len(low_sel) > n_low:
            rest = [i for i in low_sel if i != g_lo]
            if n_low == 1:
                low_sel = [g_lo]
            else:
                k = max(1, int(math.ceil(len(rest) / float(n_low - 1))))
                low_sel = [g_lo] + rest[::k][:n_low-1]

    # Emit points
    # Global first, then other picks by milepost
    def _emit(idx, kind, label, category):
        mp = float(mps[idx])
        points.append({
            'name': _nm(kind, label, mp),
            'mp': mp,
            'elev': float(elevs[idx]),
            'category': category,
            'scope': scope,
            'min_psig': None,
            'max_psig': None
        })

    _emit(g_hi, 'High', 'GLOBAL', 'auto_high')
    _emit(g_lo, 'Low', 'GLOBAL', 'auto_low')

    # Remaining picks (exclude globals), in milepost order for readability
    hi_rest = [i for i in high_sel if i != g_hi]
    lo_rest = [i for i in low_sel if i != g_lo]
    hi_rest.sort(key=lambda i: float(mps[i]))
    lo_rest.sort(key=lambda i: float(mps[i]))

    # Label bins by order along the line (not by elevation rank)
    for k, i in enumerate(hi_rest, start=1):
        _emit(i, 'High', f'BIN{k:02d}', 'auto_high')

    for k, i in enumerate(lo_rest, start=1):
        _emit(i, 'Low', f'BIN{k:02d}', 'auto_low')

    return points

# === SI-based slug friction helper (Darcy–Weisbach over remaining liquid) ===

def slug_friction_psi_SI(L_ft, D_ft, v_fts, api_gravity, viscosity_cst, eps_ft, sg_fallback=0.85):
    """
    Darcy–Weisbach dp across the *liquid slug ahead of the pig*.
    Imperial inputs, internal calc in SI, returns psi.
    """
    try:
        if L_ft <= 0 or v_fts <= 0 or D_ft <= 0:
            return 0.0
        # Imperial -> SI
        ft_to_m = 0.3048
        L = L_ft * ft_to_m
        D = D_ft * ft_to_m
        v = v_fts * ft_to_m         # ft/s -> m/s
        eps = eps_ft * ft_to_m

        # Fluid properties
        if api_gravity is None and sg_fallback is not None:
            SG = sg_fallback
        elif api_gravity is None:
            SG = 0.85
        else:
            SG = 141.5 / (api_gravity + 131.5)   # crude SG from API
        rho = 999.0 * SG                     # kg/m^3  (~water*SG)
        nu  = (viscosity_cst if viscosity_cst else 1.0) * 1.0e-6   # m^2/s   (cSt -> m^2/s)

        # Reynolds number and friction factor (Swamee-Jain)
        Re = (v * D) / max(1e-12, nu)
        if Re < 2300.0:
            f = 64.0 / max(1.0, Re)
        else:
            f = 0.25 / (math.log10((eps/(3.7*D)) + (5.74/(Re**0.9))))**2

        # Darcy–Weisbach dp (Pa) and convert to psi
        dp_Pa = f * (L / D) * 0.5 * rho * v * v
        dp_psi = dp_Pa / 6894.757
        return float(dp_psi)
    except Exception:
        return 0.0


# === Real-gas nitrogen helpers (Peng–Robinson) ===
def _f_to_k(t_f: float) -> float:
    return (t_f - 32.0) * (5.0 / 9.0) + 273.15

def _psia_to_pa(p_psia: float) -> float:
    return float(p_psia) * 6894.757

def _pa_to_psia(p_pa: float) -> float:
    return float(p_pa) / 6894.757

def z_factor_n2_pr(p_psia: float, t_f: float) -> float:
    """Return nitrogen compressibility factor Z using Peng–Robinson EOS."""
    p_pa = max(1.0, _psia_to_pa(p_psia))
    t_k = max(1.0, _f_to_k(t_f))
    R = 8.314462618  # J/mol/K

    # Peng–Robinson parameters
    kappa = 0.37464 + 1.54226 * N2_ACENTRIC - 0.26992 * (N2_ACENTRIC ** 2)
    alpha = (1.0 + kappa * (1.0 - math.sqrt(t_k / N2_CRIT_T_K))) ** 2
    a = 0.45724 * (R ** 2) * (N2_CRIT_T_K ** 2) / N2_CRIT_P_PA
    b = 0.07780 * R * N2_CRIT_T_K / N2_CRIT_P_PA
    A = a * alpha * p_pa / (R ** 2 * t_k ** 2)
    B = b * p_pa / (R * t_k)

    # PR cubic: Z^3 - (1-B)Z^2 + (A - 3B^2 - 2B)Z - (AB - B^2 - B^3) = 0
    c3 = 1.0
    c2 = -(1.0 - B)
    c1 = A - 3.0 * B * B - 2.0 * B
    c0 = -(A * B - B * B - B ** 3)

    roots = np.roots([c3, c2, c1, c0])
    # Take largest real root (gas phase)
    real_roots = [r.real for r in roots if abs(r.imag) < 1e-8]
    if not real_roots:
        return 1.0
    Z = max(real_roots)
    # Guardrails
    if not np.isfinite(Z) or Z <= 0:
        return 1.0
    return float(Z)


def n2_density_kg_m3(p_psia: float, t_f: float) -> float:
    """Nitrogen density from PR EOS: rho = P*MW/(Z*R*T)."""
    p_pa = _psia_to_pa(p_psia)
    t_k = _f_to_k(t_f)
    Z = z_factor_n2_pr(p_psia, t_f)
    R = 8.314462618
    rho = p_pa * N2_MW_KG_PER_MOL / (max(1e-12, Z) * R * t_k)
    return float(rho)


def n2_moles_from_scf(scf: float, t_std_f: float = 60.0, z_std: float = 1.0) -> float:
    """Convert SCF at standard conditions into moles."""
    # Standard conditions: 14.7 psia, 60F by convention
    p_std_pa = _psia_to_pa(ATM_PSI)
    t_std_k = _f_to_k(t_std_f)
    V_m3 = float(scf) * 0.028316846592  # ft^3 -> m^3
    R = 8.314462618
    n = p_std_pa * V_m3 / (max(1e-12, z_std) * R * t_std_k)
    return float(n)


def n2_pressure_psia_from_moles(n_mol: float, V_ft3: float, t_f: float) -> float:
    """Solve for nitrogen pressure (psia) given moles, volume, temperature using PR EOS."""
    V_m3 = max(1e-12, float(V_ft3) * 0.028316846592)
    t_k = _f_to_k(t_f)
    R = 8.314462618
    # Ideal initial guess
    p_pa = max(1.0, n_mol * R * t_k / V_m3)
    # Fixed-point iteration with Z(P)
    for _ in range(20):
        p_psia = _pa_to_psia(p_pa)
        Z = z_factor_n2_pr(p_psia, t_f)
        p_new = n_mol * R * t_k * Z / V_m3
        if abs(p_new - p_pa) / max(1.0, p_pa) < 1e-8:
            p_pa = p_new
            break
        p_pa = p_new
    return float(_pa_to_psia(p_pa))



def pressure_psig_from_moles(n_moles: float, V_ft3: float, T_F: float) -> float:
    """Convenience wrapper: nitrogen pressure (psig) from moles & volume.
    Uses the same Peng–Robinson-based real-gas helper as n2_pressure_psia_from_moles().
    """
    try:
        psia = n2_pressure_psia_from_moles(n_moles, V_ft3, T_F)
        return float(psia) - ATM_PSI
    except Exception:
        # Fallback to ideal gas if something goes sideways (should be rare)
        R = 10.7316  # (psia*ft^3)/(lbmol*R)
        T_R = float(T_F) + 459.67
        psia = (float(n_moles) * R * T_R) / max(1e-12, float(V_ft3))
        return float(psia) - ATM_PSI

def gas_friction_loss_psi_SI(L_ft: float, D_ft: float, v_fts: float, p_avg_psia: float, t_f: float, eps_ft: float) -> float:
    """Darcy–Weisbach dp (psi) for nitrogen over length L at average pressure p_avg."""
    if L_ft <= 0 or v_fts <= 0 or D_ft <= 0:
        return 0.0
    ft_to_m = 0.3048
    L = L_ft * ft_to_m
    D = D_ft * ft_to_m
    v = v_fts * ft_to_m
    eps = eps_ft * ft_to_m

    rho = n2_density_kg_m3(p_avg_psia, t_f)
    mu = N2_VISCOSITY_PA_S
    Re = rho * v * D / max(1e-12, mu)
    if Re < 2300.0:
        f = 64.0 / max(1.0, Re)
    else:
        f = 0.25 / (math.log10((eps/(3.7*D)) + (5.74/(Re**0.9))))**2

    dp_pa = f * (L / D) * 0.5 * rho * v * v
    return float(dp_pa / 6894.757)


def required_injection_pressure_psig(p_behind_pig_psig: float, L_gas_ft: float, D_ft: float, v_fts: float, eps_ft: float, t_f: float) -> tuple:
    """Return (p_inj_psig, dp_gas_psi) required to supply the pressure immediately behind the pig."""
    p_behind_pig_psia = float(p_behind_pig_psig) + ATM_PSI
    if L_gas_ft <= 0 or v_fts <= 0:
        return float(p_behind_pig_psig), 0.0
    # Iterate because density depends on pressure and dp depends on density.
    p_inj_psia = p_behind_pig_psia
    for _ in range(20):
        p_avg = 0.5 * (p_behind_pig_psia + p_inj_psia)
        dp = gas_friction_loss_psi_SI(L_gas_ft, D_ft, v_fts, p_avg, t_f, eps_ft)
        p_new = p_behind_pig_psia + dp
        if abs(p_new - p_inj_psia) / max(1.0, p_inj_psia) < 1e-8:
            p_inj_psia = p_new
            break
        p_inj_psia = p_new
    dp_gas = max(0.0, p_inj_psia - p_behind_pig_psia)
    return float(p_inj_psia - ATM_PSI), float(dp_gas)


def scfm_required_from_velocity(v_fts: float, area_ft2: float, p_inj_psig: float, p_behind_pig_psig: float, t_f: float) -> float:
    """Compute nitrogen SCFM required to sustain volumetric displacement at pig speed."""
    if v_fts <= 0 or area_ft2 <= 0:
        return 0.0
    # Use mass flow at average pressure
    p_inj_psia = float(p_inj_psig) + ATM_PSI
    p_behind_pig_psia = float(p_behind_pig_psig) + ATM_PSI
    p_avg = 0.5 * (p_inj_psia + p_behind_pig_psia)
    rho_avg = n2_density_kg_m3(p_avg, t_f)
    ft_to_m = 0.3048
    area_m2 = float(area_ft2) * (ft_to_m ** 2)
    v_m_s = float(v_fts) * ft_to_m
    m_dot = rho_avg * v_m_s * area_m2  # kg/s

    # Standard density at 14.7 psia, 60F
    rho_std = n2_density_kg_m3(ATM_PSI, 60.0)
    q_std_m3_s = m_dot / max(1e-12, rho_std)
    scfm = q_std_m3_s * 35.3146667 * 60.0
    return float(scfm)


def run_simulation(dialog_root, inputs, purge_mileposts, elevations, system_mileposts, system_elevations):
    # Work on a copy so we don't mutate inputs passed to export_results
    cfg = dict(inputs)
    # Default temperature (F) used by nitrogen/gas property helpers
    cfg.setdefault('temperature_f', float(N2_TEMP_F_DEFAULT))


    # ---- Simulation guardrails ----
    # These prevent accidental "runaway" scenarios (very large profiles, tiny velocities, etc.)
    sim_t0 = time.monotonic()
    try:
        max_wall_time_s = float(cfg.get('max_wall_time_s', 180.0) or 180.0)
    except Exception:
        max_wall_time_s = 180.0
    try:
        # Very large default; set smaller if you'd like to enforce a hard operational cap.
        max_elapsed_time_hr = float(cfg.get('max_elapsed_time_hr', 1.0e6) or 1.0e6)
    except Exception:
        max_elapsed_time_hr = 1.0e6
    try:
        max_sim_steps = int(cfg.get('max_sim_steps', 0) or 0)  # 0 => no extra cap beyond profile size
    except Exception:
        max_sim_steps = 0
    stop_reason = None


    # ---- normalize numeric inputs to floats (robust to Tk string values) ----
    def _to_float(val, default=None):
        try:
            return float(val)
        except (TypeError, ValueError):
            return default

    _numeric_fields = [
        'pipe_wt','viscosity_cst','api_gravity','max_nitrogen_pressure','max_drive_pressure',
        'exit_pressure_run','exit_pressure_end','purge_end_mp','purge_start_mp',
        'roughness_num','max_n2_rate_scfm','max_pig_speed','min_pig_speed',
        'n2_end_pressure','resolution','system_end_mp','target_pig_speed','slack_pressure_psig','maop_psig','max_outlet_flow_bph'
    ]
    for _k in _numeric_fields:
        if _k in cfg:
            _v = _to_float(cfg[_k], None)
            if _v is not None:
                cfg[_k] = _v

    # Resolve NPS
    try:
        nps_key = resolve_nps_key(cfg['nps'])
    except KeyError as e:
        raise ValueError(str(e))
    pipe_od = nps_data[nps_key]["OD_in"]
    pipe_id = pipe_od - 2 * cfg['pipe_wt']
    if pipe_id <= 0:
        raise ValueError(f"Computed ID <= 0. Check NPS {nps_key} and wall {cfg['pipe_wt']}.")
    pipe_diameter = pipe_id / 12.0  # ft
    area = np.pi * (pipe_diameter / 2) ** 2

    # Optional max outlet flow constraint (e.g., tankage acceptance limit)
    max_outlet_flow_bph = cfg.get('max_outlet_flow_bph', None)
    v_flow_cap = math.inf
    if max_outlet_flow_bph is not None:
        try:
            max_outlet_flow_bph = float(max_outlet_flow_bph)
        except Exception:
            max_outlet_flow_bph = None
    if max_outlet_flow_bph is not None and max_outlet_flow_bph > 0:
        q_cap_ft3_s = max_outlet_flow_bph * 5.614583333333333 / 3600.0
        v_flow_cap = q_cap_ft3_s / area  # ft/s
        cfg['max_pig_speed_from_max_outlet_flow_mph'] = v_flow_cap / 1.46667

    roughness = roughness_data[int(cfg['roughness_num'])]['roughness_ft']
    
    # Fluid properties
    if int(cfg['fluid_num']) == 3:
        specific_gravity = 141.5 / (131.5 + cfg['api_gravity'])  # SG from API
        viscosity_cst = cfg['viscosity_cst']
    else:
        specific_gravity = fluid_data[int(cfg['fluid_num'])]['sg']
        viscosity_cst = fluid_data[int(cfg['fluid_num'])]['viscosity_cst']

    api_for_dp = cfg['api_gravity'] if int(cfg['fluid_num']) == 3 else None
    # Use weight density (lb/ft^3) for hydrostatic, convert to mass density when needed
    gamma = specific_gravity * 62.4
    fluid_density = gamma  # keep legacy name for compatibility
    viscosity = (gamma / 32.174) * (viscosity_cst * 1.076e-5)  # dynamic mu (lb*s/ft^2)

    purge_length = (cfg['purge_end_mp'] - cfg['purge_start_mp']) * 5280.0
    total_volume_ft3 = area * purge_length
    total_volume_scf = total_volume_ft3 / SCF_TO_FT3

    # --- Nitrogen temperature & real-gas Z-factor ---
    # Per user convention we assume pipeline/gas temperature is approximately ground temperature.
    n2_temp_f = float(cfg.get('n2_temp_f', N2_TEMP_F_DEFAULT) or N2_TEMP_F_DEFAULT)

    # Compute the automatic nitrogen cut‑off volume (SCF @ 14.7 psia, 60F) based on
    # the desired end pressure *throughout the purged volume*.
    # Real-gas correction: V_std = V_pipe * (P_end/P_std) * (T_std/T_pipe) * (Z_std/Z_end)
    P_end_abs = float(cfg['n2_end_pressure']) + ATM_PSI
    Z_end = z_factor_n2_pr(P_end_abs, n2_temp_f)
    Z_std = z_factor_n2_pr(ATM_PSI, 60.0)
    T_std_R = 60.0 + 459.67
    T_pipe_R = n2_temp_f + 459.67
    cutoff_volume = total_volume_ft3 * (P_end_abs / ATM_PSI) * (T_std_R / max(1e-12, T_pipe_R)) * (Z_std / max(1e-12, Z_end))
    logging.info(
        f"Total pipe volume: {total_volume_ft3:.0f} ft³, {total_volume_scf:.0f} SCF, "
        f"Cutoff (N2) volume: {cutoff_volume:.0f} SCF"
    )
    
    n_points = len(purge_mileposts)
    distances = purge_mileposts * 5280.0

    # Static head: positive means uphill (pressure you must overcome), negative downhill (assist)
    cs = CubicSpline(system_mileposts, system_elevations)
    h_exit = cs(cfg['system_end_mp'])
    head_losses = fluid_density * (h_exit - elevations) / 144.0  # psi

    # --- Monitoring setup (pressure at critical points) ---
    slack_threshold_psig = float(cfg.get('slack_pressure_psig', 50.0))
    maop_psig = cfg.get('maop_psig', None)
    try:
        maop_psig = float(maop_psig) if maop_psig is not None else None
    except Exception:
        maop_psig = None

    
    # Auto-monitor extrema (high/low points) along the modeled hydraulic system / purge segment
    monitor_points = []
    if bool(cfg.get('auto_monitor_enabled', True)):
        try:
            n_hi = int(cfg.get('auto_monitor_n_high', 10))
        except Exception:
            n_hi = 10
        try:
            n_lo = int(cfg.get('auto_monitor_n_low', 10))
        except Exception:
            n_lo = 10
        try:
            spacing_mi = float(cfg.get('auto_monitor_min_spacing_mi', 0.5))
        except Exception:
            spacing_mi = 0.5

        # Auto-monitor selection strategy:
        #   mode='binned' picks per-section highs/lows across the full profile (recommended)
        #   mode='local_extrema' picks the top-N local turning points
        mode = str(cfg.get('auto_monitor_mode', 'binned')).strip().lower()
        if mode not in ('binned', 'local_extrema', 'local', 'peaks'):
            mode = 'binned'
        try:
            n_bins = int(cfg.get('auto_monitor_bins', 10))
        except Exception:
            n_bins = 10

        scope = str(cfg.get('auto_monitor_scope', 'system')).strip().lower()
        if scope not in ('system', 'purge', 'both'):
            scope = 'system'

        if scope in ('system', 'both'):
            monitor_points.extend(build_auto_monitor_points(
                system_mileposts, system_elevations,
                n_high=n_hi, n_low=n_lo, min_spacing_mi=spacing_mi,
                name_prefix="SYS ", n_bins=n_bins, mode=mode
            ))
        if scope in ('purge', 'both'):
            monitor_points.extend(build_auto_monitor_points(
                purge_mileposts, elevations,
                n_high=n_hi, n_low=n_lo, min_spacing_mi=spacing_mi,
                name_prefix="PURGE ", n_bins=n_bins, mode=mode
            ))

    # Intermediate pump stations: accept multiple mileposts; treat as monitored nodes (not hydraulic boundaries).
    ips_mps = []
    if bool(cfg.get('has_ips', False)):
        ips_mps = parse_milepost_list(cfg.get('ips_mps', None) or cfg.get('ips_mp', None))
        # keep only those inside modeled segment
        ips_mps = [float(m) for m in ips_mps if cfg['purge_start_mp'] < float(m) < cfg['system_end_mp']]
        ips_mps = sorted(list(dict.fromkeys(ips_mps)))
        for k, mp in enumerate(ips_mps, start=1):
            try:
                elev_mp = float(cs(mp))
            except Exception:
                elev_mp = float(np.interp(mp, purge_mileposts, elevations))
            monitor_points.append({
                'name': f'IPS#{k} @ MP {mp:.3f}',
                'mp': float(mp),
                'elev': elev_mp,
                'category': 'pump_station'
            })
    cfg['ips_mps'] = ips_mps

    # Attach monitor rules: 
    # - MAOP rule applies at all monitor points (liquid or gas)
    # - Slack / minimum-pressure rule applies ONLY when the point is ahead of the pig (liquid-filled)
    min_pump_suction_psig = cfg.get('min_pump_suction_pressure', None)
    try:
        min_pump_suction_psig = float(min_pump_suction_psig) if min_pump_suction_psig is not None else None
    except Exception:
        min_pump_suction_psig = None

    for p in monitor_points:
        p.setdefault('category', 'user')
        p['min_psig'] = None
        p['max_psig'] = None
        if maop_psig is not None:
            p['max_psig'] = maop_psig

        if p.get('category') in ('auto_high', 'pump_station'):
            p['min_psig'] = slack_threshold_psig
            if p.get('category') == 'pump_station' and min_pump_suction_psig is not None:
                p['min_psig'] = max(p['min_psig'], min_pump_suction_psig)

        if p.get('category') == 'auto_low':
            # lows: only MAOP typically relevant; slack doesn't apply automatically
            p['min_psig'] = None

    # Ensure unique names
    _seen = set()
    for p in monitor_points:
        base = p['name']
        name = base
        n = 2
        while name in _seen:
            name = f"{base} ({n})"
            n += 1
        p['name'] = name
        _seen.add(name)

    monitor_pressures = {p['name']: np.full(n_points, np.nan, dtype=np.float64) for p in monitor_points}
    # Hypothetical blocked-in static state monitoring (psig) at the same monitor points
    monitor_pressures_static = {p['name']: np.full(n_points, np.nan, dtype=np.float64) for p in monitor_points}

    # Static block-in diagnostics (computed each step)
    static_hgl_ft = np.full(n_points, np.nan, dtype=np.float64)
    static_max_any_psig = np.full(n_points, np.nan, dtype=np.float64)
    static_max_any_mp = np.full(n_points, np.nan, dtype=np.float64)
    static_max_any_elev_ft = np.full(n_points, np.nan, dtype=np.float64)
    static_max_any_phase = np.array([''] * n_points, dtype=object)

    static_violations = []
    monitor_violations = []

    elapsed_times = np.zeros(n_points, dtype=np.float64)
    # Pressure immediately behind the pig (psig). Historically this was called
    # "drive pressure" in the program, but "pressure behind pig" is clearer.
    behind_pig_pressures = np.zeros(n_points, dtype=np.float64)
    # Inlet injection pressure at purge start (psig)
    injection_pressures = np.zeros(n_points, dtype=np.float64)
    # Gas friction dp from injection point -> pig (psi)
    gas_dp_losses = np.zeros(n_points, dtype=np.float64)

    # Differential pressure diagnostic: injection pressure - exit pressure (psi)
    differential_pressures = np.zeros(n_points, dtype=np.float64)
    # Miles remaining to system outlet (miles)
    miles_to_outlet = np.zeros(n_points, dtype=np.float64)

    friction_losses = np.zeros(n_points, dtype=np.float64)
    exit_pressures = np.zeros(n_points, dtype=np.float64)
    # Diagnostics for endpoint pressure constraint scheme (optional)
    exit_pressures_programmed = np.full(n_points, np.nan, dtype=np.float64)
    exit_pressures_lb = np.full(n_points, np.nan, dtype=np.float64)
    exit_pressures_ub = np.full(n_points, np.nan, dtype=np.float64)
    injection_rates = np.zeros(n_points, dtype=np.float64)
    cumulative_n2 = np.zeros(n_points, dtype=np.float64)
    pig_speeds = np.zeros(n_points, dtype=np.float64)
    exceed_points = []

    # --- Intermediate Pump Station (IPS) runtime diagnostics ---
    hydraulic_boundary_mps = np.full(n_points, np.nan, dtype=np.float64)          # MP of the downstream boundary used in upstream math
    hydraulic_boundary_pressures = np.full(n_points, np.nan, dtype=np.float64)    # Pressure at that boundary (psig)
    head_to_boundary = np.full(n_points, np.nan, dtype=np.float64)                # Static head from pig to boundary (psi, + if boundary is higher)
    ips_active_mp = np.full(n_points, np.nan, dtype=np.float64)                   # Active IPS MP when check is sealed, else NaN
    ips_suction_psig = np.full(n_points, np.nan, dtype=np.float64)                # Suction pressure used/assumed at active IPS (psig)
    ips_discharge_req_psig = np.full(n_points, np.nan, dtype=np.float64)          # Discharge pressure required to satisfy endpoint BC (psig)
    ips_pump_dp_req_psi = np.full(n_points, np.nan, dtype=np.float64)             # Pump ΔP required (psi)
    ips_flow_bph = np.full(n_points, np.nan, dtype=np.float64)                    # Mainline flow (bph)
    ips_state = np.array([""] * n_points, dtype=object)                           # Text state for debugging / exports
    
    cumulative_n2[0] = 0.0
    cumulative_distance = 0.0
    last_valid_i = n_points - 1
    v_max = cfg['max_pig_speed'] * 1.46667  # ft/s
    v_target = float(cfg.get('target_pig_speed', cfg['max_pig_speed']) or cfg['max_pig_speed']) * 1.46667

    # Apply max outlet-flow-derived velocity cap, if configured
    if v_flow_cap != math.inf:
        v_max = min(v_max, v_flow_cap)
        v_target = min(v_target, v_flow_cap)
    max_inj_psig = float(cfg.get('max_nitrogen_pressure', cfg.get('max_drive_pressure', 400.0)) or 400.0)  # inlet cap

    # Max N2 rate: None -> unlimited ("auto")
    max_rate_in = cfg.get('max_n2_rate_scfm', None)
    try:
        max_rate = float(max_rate_in) if (max_rate_in is not None and float(max_rate_in) > 0) else math.inf
    except Exception:
        max_rate = math.inf

    # Optional nitrogen cutoff (SCF) -- stop injection and coast using stored gas
    # Determine nitrogen cut‑off volume.  If the user specified a
    # manual cut‑off (n2_cutoff_scf) use it; otherwise fall back to
    # the automatically computed cutoff_volume.  A non‑positive
    # user‑provided value is ignored.
    n2_cutoff = cfg.get('n2_cutoff_scf', None)
    if n2_cutoff is not None:
        try:
            n2_cutoff = float(n2_cutoff)
            if n2_cutoff <= 0:
                n2_cutoff = None
        except Exception:
            n2_cutoff = None
    # Use automatic cut‑off if no valid manual value is provided
    cutoff_value = n2_cutoff if n2_cutoff is not None else cutoff_volume
    coast_mode = False
    n2_moles_total = None   # mol (at standard, injected)
    V_gas_ft3 = None        # ft^3 (geometric volume behind pig)

    # --- IPS / Pump Station runtime parameters (global, applied to each station in ips_mps) ---
    try:
        ips_shutdown_dist = float(cfg.get('ips_shutdown_dist', 0.0) or 0.0)
    except Exception:
        ips_shutdown_dist = 0.0
    try:
        min_pump_suction_pressure = float(cfg.get('min_pump_suction_pressure', 0.0) or 0.0)
    except Exception:
        min_pump_suction_pressure = 0.0
    min_pump_flow_bph = cfg.get('min_pump_flow_bph', None)
    try:
        min_pump_flow_bph = float(min_pump_flow_bph) if (min_pump_flow_bph is not None and float(min_pump_flow_bph) > 0) else None
    except Exception:
        min_pump_flow_bph = None
    try:
        ips_check_dp_psi = float(cfg.get('ips_check_dp_psi', cfg.get('ips_check_dp', 5.0)) or 5.0)
    except Exception:
        ips_check_dp_psi = 5.0

    def _bph_from_v(v_fts, area_ft2):
        # bph = (ft^3/s * 3600 s/hr) / (5.614583 ft^3/bbl)
        return (max(0.0, float(v_fts)) * float(area_ft2) * 3600.0) / 5.614583

    def _v_from_bph(bph, area_ft2):
        if area_ft2 <= 0:
            return 0.0
        return (float(bph) * 5.614583) / (3600.0 * float(area_ft2))

    # Persist previous endpoint BC for optional ramp-rate limiting
    P_end_prev = None

    max_steps = n_points - 1
    if max_sim_steps and max_sim_steps > 0:
        max_steps = min(max_steps, int(max_sim_steps))

    for i in range(max_steps):
        segment_length_ft = max(1e-6, distances[i + 1] - distances[i])
        pig_mp_now = float(purge_mileposts[i])
        elev_pig = float(elevations[i])
        # Wall-clock guard (prevents "hangs" on massive inputs)
        if (time.monotonic() - sim_t0) > max_wall_time_s:
            stop_reason = f"TIMEOUT>{max_wall_time_s:.0f}s"
            last_valid_i = max(0, i - 1)
            logging.warning(f"Simulation timeout reached at i={i}, MP={pig_mp_now:.3f}. Stopping early.")
            break


        # Endpoint boundary condition (system end) using selected behavior
        P_end_prog = float(target_exit_pressure(pig_mp_now, cfg, cfg['purge_start_mp'], cfg['purge_end_mp'], cfg['throttle_down_miles']))
        P_end = float(P_end_prog)
        exit_pressures_programmed[i] = float(P_end_prog)

        # Gas length behind pig (from purge start)
        L_gas_ft = max(0.0, (pig_mp_now - cfg['purge_start_mp']) * 5280.0)

        # --- Available pressure behind pig in coast mode (adiabatic-ish expansion at fixed moles) ---
        P_trap_psig = 0.0
        if coast_mode and cutoff_value is not None:
            if n2_moles_total is None:
                n2_moles_total = n2_moles_from_scf(cutoff_value)
            # Ensure we have a gas volume estimate
            if V_gas_ft3 is None:
                V_gas_ft3 = max(1e-6, area * max(1e-6, L_gas_ft))
            P_trap_psig = pressure_psig_from_moles(n2_moles_total, V_gas_ft3, cfg.get('temperature_f', 60.0))
        else:
            P_trap_psig = 0.0

        # --- Helper: solve max feasible velocity for a given downstream boundary (MP, pressure) ---
        def solve_boundary(boundary_mp, P_boundary, elev_boundary, label=""):
            slug_length_ft_local = max(0.0, (float(boundary_mp) - pig_mp_now) * 5280.0)
            head_local = float(fluid_density) * (float(elev_boundary) - elev_pig) / 144.0

            def evaluate(v_try, enforce_caps=True):
                v_try = float(v_try)
                fr_liq = slug_friction_psi_SI(slug_length_ft_local, pipe_diameter, v_try, api_for_dp, viscosity_cst, roughness, sg_fallback=specific_gravity)
                P_behind_pig_req = max(float(P_boundary), float(P_boundary) + head_local + fr_liq)

                if coast_mode and cutoff_value is not None:
                    cap = float(P_trap_psig)
                    ok = (P_behind_pig_req <= cap + 1e-6) if enforce_caps else True
                    P_behind_pig_out = min(cap, P_behind_pig_req) if enforce_caps else P_behind_pig_req
                    P_inj_out = P_behind_pig_out
                    dp_gas = 0.0
                    q_use = 0.0
                    return ok, P_behind_pig_req, P_behind_pig_out, P_inj_out, dp_gas, q_use, fr_liq

                # Injection mode
                P_inj_req, dp_gas = required_injection_pressure_psig(P_behind_pig_req, L_gas_ft, pipe_diameter, v_try, roughness, cfg.get('temperature_f', 60.0))
                q_required = scfm_required_from_velocity(v_try, area, P_inj_req, P_behind_pig_req, cfg.get('temperature_f', 60.0))

                ok = True
                if enforce_caps:
                    ok = (P_inj_req <= max_inj_psig + 1e-6) and (q_required <= max_rate + 1e-6)

                P_behind_pig_out = P_behind_pig_req
                P_inj_out = P_inj_req
                dp_gas = float(dp_gas)
                q_use = q_required
                return ok, P_behind_pig_req, P_behind_pig_out, P_inj_out, dp_gas, q_use, fr_liq

            # Quick stall check
            ok0, P_req0, P_out0, P_inj0, dp_gas0, q0, fr0 = evaluate(1e-6, enforce_caps=True)
            if not ok0:
                return {
                    'feasible': False, 'v': 0.0, 'P_behind_req': P_req0, 'P_behind_out': P_out0,
                    'P_inj': P_inj0, 'dp_gas': dp_gas0, 'q_use': q0, 'fr_liq': fr0,
                    'head': head_local, 'slug_length_ft': slug_length_ft_local
                }

            # Upper velocity bound
            v_hi = v_max if (coast_mode and cutoff_value is not None) else min(v_max, v_target)
            ok_hi, P_req_hi, P_out_hi, P_inj_hi, dp_gas_hi, q_hi, fr_hi = evaluate(v_hi, enforce_caps=True)

            if ok_hi:
                return {
                    'feasible': True, 'v': float(v_hi), 'P_behind_req': P_req_hi, 'P_behind_out': P_out_hi,
                    'P_inj': P_inj_hi, 'dp_gas': dp_gas_hi, 'q_use': q_hi, 'fr_liq': fr_hi,
                    'head': head_local, 'slug_length_ft': slug_length_ft_local
                }

            # Bisection for maximum feasible v
            lo, hi = 0.0, float(v_hi)
            best = 0.0
            best_pack = (P_req0, P_out0, P_inj0, dp_gas0, q0, fr0)
            for _ in range(40):
                if (hi - lo) < 1e-6:
                    break
                mid = 0.5 * (lo + hi)
                ok_mid, P_req_mid, P_out_mid, P_inj_mid, dp_gas_mid, q_mid, fr_mid = evaluate(mid, enforce_caps=True)
                if ok_mid:
                    lo = mid
                    best = mid
                    best_pack = (P_req_mid, P_out_mid, P_inj_mid, dp_gas_mid, q_mid, fr_mid)
                else:
                    hi = mid

            P_req_b, P_out_b, P_inj_b, dp_gas_b, q_b, fr_b = best_pack
            return {
                'feasible': True, 'v': float(best), 'P_behind_req': P_req_b, 'P_behind_out': P_out_b,
                'P_inj': P_inj_b, 'dp_gas': dp_gas_b, 'q_use': q_b, 'fr_liq': fr_b,
                'head': head_local, 'slug_length_ft': slug_length_ft_local
            }

        # --- Baseline: pumps OFF / flow-through at IPS (boundary at system end) ---
        endpoint_mode = str(cfg.get('endpoint_constraint_mode', 'none')).strip().lower()
        exit_min_clamp = cfg.get('exit_pressure_min_clamp', None)
        exit_max_clamp = cfg.get('exit_pressure_max_clamp', None)
        exit_ramp = cfg.get('exit_pressure_ramp_psi_per_hr', None)
        try:
            exit_min_clamp = float(exit_min_clamp) if exit_min_clamp is not None else None
        except Exception:
            exit_min_clamp = None
        try:
            exit_max_clamp = float(exit_max_clamp) if exit_max_clamp is not None else None
        except Exception:
            exit_max_clamp = None
        try:
            exit_ramp = float(exit_ramp) if exit_ramp is not None else None
        except Exception:
            exit_ramp = None

        def _exit_bounds_for_monitors(v_fts: float):
            """Compute allowable endpoint pressure bounds (lb, ub) such that monitored liquid points ahead of pig
            do not violate slack/min and MAOP/max limits, for the IPS-OFF / flow-through case."""
            lb = -1.0e9
            ub =  1.0e9
            v_fts = float(max(0.0, v_fts))

            for p in monitor_points:
                try:
                    mp_p = float(p.get('mp'))
                except Exception:
                    continue
                if mp_p < pig_mp_now - 1e-9:
                    continue  # behind pig => gas in this model

                # Reference boundary is the *system end* (IPS off / check open)
                try:
                    elev_p = float(p.get('elev', cs(mp_p)))
                except Exception:
                    elev_p = float(np.interp(mp_p, purge_mileposts, elevations))

                L_seg_ft = max(0.0, (float(cfg['system_end_mp']) - mp_p) * 5280.0)
                fr_seg = slug_friction_psi_SI(L_seg_ft, pipe_diameter, v_fts, api_for_dp, viscosity_cst, roughness, sg_fallback=specific_gravity) if v_fts > 1e-10 else 0.0
                head_seg = float(fluid_density) * (float(h_exit) - elev_p) / 144.0
                coeff = float(head_seg) + float(fr_seg)  # P(point) = P_end + coeff

                # Upper bounds (MAOP)
                max_lim = p.get('max_psig', None)
                if max_lim is not None:
                    try:
                        ub = min(ub, float(max_lim) - coeff)
                    except Exception:
                        pass

                # Lower bounds: per-point min (if provided) and global slack threshold (liquid section only)
                min_lim = p.get('min_psig', None)
                if min_lim is not None:
                    try:
                        lb = max(lb, float(min_lim) - coeff)
                    except Exception:
                        pass

                lb = max(lb, float(slack_threshold_psig) - coeff)

            # Manual clamps
            if exit_min_clamp is not None:
                lb = max(lb, float(exit_min_clamp))
            if exit_max_clamp is not None:
                ub = min(ub, float(exit_max_clamp))

            return lb, ub

        # Start from the programmed BC, then optionally clamp & ramp it.
        P_end_use = float(P_end)

        # Monitor-based clamp uses the IPS-OFF (endpoint boundary) regime only.
        if endpoint_mode == "clamp_to_monitors":
            P_try = float(P_end_use)
            sol_tmp = None
            for _ in range(4):
                sol_tmp = solve_boundary(cfg['system_end_mp'], P_try, h_exit, label="OFF_CLAMP")
                if not sol_tmp.get('feasible', False):
                    break
                v_tmp = float(sol_tmp.get('v', 0.0))
                lb, ub = _exit_bounds_for_monitors(v_tmp)
                exit_pressures_lb[i] = lb
                exit_pressures_ub[i] = ub

                if lb > ub:
                    logging.warning(
                        f"Endpoint constraint infeasible at pig MP {pig_mp_now:.3f}: "
                        f"LB={lb:.2f} psig > UB={ub:.2f} psig. Clamping to UB to protect MAOP."
                    )
                    P_new = min(float(P_end_prog), float(ub))
                else:
                    P_new = min(max(float(P_end_prog), float(lb)), float(ub))

                if abs(P_new - P_try) < 0.05:
                    P_try = float(P_new)
                    break
                P_try = float(P_new)

            P_end_use = float(P_try)

        # Apply optional ramp-rate limit (psi/hr). This is applied after the clamp.
        if (exit_ramp is not None) and (P_end_prev is not None):
            # Estimate dt using a first-pass solution at current P_end_use.
            sol_dt = solve_boundary(cfg['system_end_mp'], P_end_use, h_exit, label="OFF_DT")
            if sol_dt.get('feasible', False):
                v_dt = max(1e-9, float(sol_dt.get('v', 0.0)))
                dt_est_hr = float(segment_length_ft) / v_dt / 3600.0
                max_delta = float(exit_ramp) * max(0.0, dt_est_hr)
                P_end_use = min(max(P_end_use, float(P_end_prev) - max_delta), float(P_end_prev) + max_delta)

        # Final OFF solution at the selected endpoint pressure
        P_end = float(P_end_use)
        sol_off = solve_boundary(cfg['system_end_mp'], P_end, h_exit, label="OFF")
        v_off = float(sol_off['v']) if sol_off.get('feasible', False) else 0.0

        # Persist endpoint BC for next iteration (for ramp limiting)
        P_end_prev = float(P_end)

        # Default to OFF solution unless an IPS is both available and actually acting as a boundary
        use_ips = False
        active_ips = None
        ips_reason = "OFF"
        sol_use = sol_off
        P_boundary_use = P_end
        boundary_mp_use = float(cfg['system_end_mp'])
        elev_boundary_use = float(h_exit)

        # Find the next downstream IPS ahead of the pig (closest MP > pig MP)
        next_ips = None
        if cfg.get('has_ips', False) and isinstance(ips_mps, (list, tuple)) and len(ips_mps) > 0:
            for _mp in ips_mps:
                if float(_mp) > pig_mp_now:
                    next_ips = float(_mp)
                    break

        in_shutdown_zone = False
        if next_ips is not None and ips_shutdown_dist > 0.0:
            in_shutdown_zone = (pig_mp_now >= (next_ips - ips_shutdown_dist))

        # --- Decide whether the IPS actually becomes the upstream hydraulic boundary ---
        if next_ips is not None and (not in_shutdown_zone):
            # If the line can already hit target without the station (drive pressure sufficient), treat as flow-through
            if (not sol_off['feasible']) or (v_off < (v_target * 0.995)):
                # Attempt IPS-on: treat the IPS suction as the downstream boundary for upstream math
                h_ips = float(cs(next_ips))
                P_suction_set = max(float(min_pump_suction_pressure), float(slack_threshold_psig))
                sol_on = solve_boundary(next_ips, P_suction_set, h_ips, label="ON")
                v_on = float(sol_on['v']) if sol_on['feasible'] else 0.0

                # Compute downstream discharge requirement at the actual flow (for check sealing / whether station is a true boundary)
                if sol_on['feasible'] and (next_ips < float(cfg['system_end_mp']) - 1e-9):
                    L_down_ft = max(0.0, (float(cfg['system_end_mp']) - next_ips) * 5280.0)
                    head_down = float(fluid_density) * (float(h_exit) - h_ips) / 144.0
                    fr_down = slug_friction_psi_SI(L_down_ft, pipe_diameter, v_on, api_for_dp, viscosity_cst, roughness, sg_fallback=specific_gravity)
                    P_discharge_req = max(float(P_end), float(P_end) + head_down + fr_down)
                    dp_pump_req = float(P_discharge_req) - float(P_suction_set)
                else:
                    P_discharge_req = float(P_end)
                    dp_pump_req = 0.0

                # Minimum-flow constraint (if provided): station only "in use" when mainline flow meets/exceeds this.
                bph_on = _bph_from_v(v_on, area)
                minflow_ok = True if (min_pump_flow_bph is None) else (bph_on >= float(min_pump_flow_bph) - 1e-9)

                # Check valve seal logic: if pump ΔP is too small, assume check never fully seals -> not a boundary.
                seals_check = (dp_pump_req > float(ips_check_dp_psi) + 1e-9)

                if sol_on['feasible'] and minflow_ok and seals_check and (v_on > v_off + 1e-9):
                    # IPS is acting as a hydraulic boundary for upstream math
                    use_ips = True
                    active_ips = next_ips
                    sol_use = sol_on
                    P_boundary_use = float(P_suction_set)
                    boundary_mp_use = float(next_ips)
                    elev_boundary_use = float(h_ips)
                    ips_reason = "ON" if (v_on >= v_target * 0.995) else "ON_FLOW_SHORTFALL"

                    ips_active_mp[i] = float(next_ips)
                    ips_suction_psig[i] = float(P_suction_set)
                    ips_discharge_req_psig[i] = float(P_discharge_req)
                    ips_pump_dp_req_psi[i] = float(dp_pump_req)
                    ips_flow_bph[i] = float(bph_on)
                    ips_state[i] = ips_reason
                else:
                    if in_shutdown_zone:
                        ips_state[i] = "SHUTDOWN_ZONE"
                    elif not sol_on['feasible']:
                        ips_state[i] = "ON_STALL"
                    elif not minflow_ok:
                        ips_state[i] = "ON_BELOW_MINFLOW"
                    elif not seals_check:
                        ips_state[i] = "CHECK_NOT_SEALED"
                    else:
                        ips_state[i] = "NO_BENEFIT"
            else:
                ips_state[i] = "DRIVE_SUFFICIENT"
        elif next_ips is not None and in_shutdown_zone:
            ips_state[i] = "SHUTDOWN_ZONE"
        else:
            ips_state[i] = "NO_STATION_AHEAD"

        # --- Use selected solution to update primary arrays ---
        v = float(sol_use['v']) if sol_use['feasible'] else 0.0
        P_behind_pig_req = float(sol_use['P_behind_req'])
        P_behind_pig_out = float(sol_use['P_behind_out'])
        P_inj_out = float(sol_use['P_inj'])
        dp_gas = float(sol_use['dp_gas'])
        q_use = float(sol_use['q_use'])
        fr_use = float(sol_use['fr_liq'])
        head_use = float(sol_use['head'])

        injection_pressures[i] = P_inj_out
        behind_pig_pressures[i] = P_behind_pig_out
        differential_pressures[i] = P_inj_out - P_end
        injection_rates[i] = 0.0 if (coast_mode and cutoff_value is not None) else q_use
        pig_speeds[i] = v / 1.46667
        friction_losses[i] = fr_use
        exit_pressures[i] = float(P_end)
        miles_to_outlet[i] = float(cfg['system_end_mp'] - pig_mp_now)

        hydraulic_boundary_mps[i] = float(boundary_mp_use)
        hydraulic_boundary_pressures[i] = float(P_boundary_use)
        head_to_boundary[i] = float(head_use)
        ips_flow_bph[i] = _bph_from_v(v, area) if not use_ips else ips_flow_bph[i]
        if use_ips:
            ips_state[i] = ips_state[i]  # already set above
        else:
            # Keep flow diagnostic even when station isn't in use (helpful)
            ips_flow_bph[i] = _bph_from_v(v, area)

        
        # --- Hypothetical blocked-in static check (pressurized equilibrium) ---
        # Intent: "If we shut the outlet valve right now (flow -> 0), would we exceed MAOP anywhere?"
        #
        # Key physics:
        #   - With flow stopped, friction terms go to ~0. The liquid section becomes purely hydrostatic.
        #   - But it is NOT "zero-pressure" hydrostatic; it is anchored to the (pressurized) gas boundary behind the pig.
        #   - We approximate the liquid-side interface pressure as the current/trapped nitrogen pressure behind the pig.
        #
        # For any liquid-filled point (ahead of pig):
        #   P_static(point) = P_gas_static + rho*(elev_pig - elev_point)/144
        # This is equivalent to a constant static HGL:
        #   HGL_static = elev_pig + P_gas_static*144/rho
        hgl_static_liq_ft = np.nan
        gas_static_psig = np.nan
        if bool(cfg.get('static_block_in_enabled', True)):
            # Pig elevation (ft)
            try:
                elev_pig = float(cs(pig_mp_now))
            except Exception:
                elev_pig = float(np.interp(pig_mp_now, system_mileposts, system_elevations))
        
            # Estimate trapped gas pressure behind pig (psig) from moles + volume (EOS helper),
            # and do NOT allow it to be below the current computed behind-pig pressure (conservative).
            try:
                n2_scf_now = float(cumulative_n2[i])
            except Exception:
                n2_scf_now = 0.0
            n2_moles_now = n2_moles_from_scf(max(0.0, n2_scf_now))
            L_gas_ft_static = max(1e-6, (pig_mp_now - float(cfg['purge_start_mp'])) * 5280.0)
            V_gas_ft3_static = max(1e-6, float(area) * float(L_gas_ft_static))
            try:
                P_trap_now = float(pressure_psig_from_moles(n2_moles_now, V_gas_ft3_static, cfg.get('temperature_f', 60.0)))
            except Exception:
                P_trap_now = float(P_behind_pig_out)
            P_gas_static = float(max(float(P_behind_pig_out), float(P_trap_now)))
            P_gas_static = float(max(0.0, P_gas_static))
        
            gas_static_psig = P_gas_static
            hgl_static_liq_ft = float(elev_pig + (P_gas_static * 144.0) / float(fluid_density))
            static_hgl_ft[i] = hgl_static_liq_ft
        
            # Worst-case static pressure anywhere (liquid via hydrostatic from HGL; gas uniform)
            try:
                sys_mps_arr = np.asarray(system_mileposts, dtype=float)
                sys_elev_arr = np.asarray(system_elevations, dtype=float)
                P_static_liq_all = (hgl_static_liq_ft - sys_elev_arr) * float(fluid_density) / 144.0
                liq_mask = sys_mps_arr >= pig_mp_now - 1e-9
                max_liq = np.nan
                max_liq_mp = np.nan
                max_liq_elev = np.nan
                if np.any(liq_mask):
                    masked = np.where(liq_mask, P_static_liq_all, -np.inf)
                    j = int(np.nanargmax(masked))
                    max_liq = float(masked[j])
                    max_liq_mp = float(sys_mps_arr[j])
                    max_liq_elev = float(sys_elev_arr[j])
            except Exception:
                max_liq = np.nan
                max_liq_mp = np.nan
                max_liq_elev = np.nan
        
            # Gas max location: represent at purge start (uniform in gas section)
            try:
                elev_inj = float(cs(float(cfg['purge_start_mp'])))
            except Exception:
                elev_inj = float(np.interp(float(cfg['purge_start_mp']), system_mileposts, system_elevations))
        
            gas_max = float(P_gas_static)
            gas_max_mp = float(cfg['purge_start_mp'])
            gas_max_elev = float(elev_inj)
        
            # Combine phases for "worst anywhere"
            if np.isfinite(max_liq) and (float(max_liq) >= float(gas_max)):
                static_max_any_psig[i] = float(max_liq)
                static_max_any_mp[i] = float(max_liq_mp)
                static_max_any_elev_ft[i] = float(max_liq_elev)
                static_max_any_phase[i] = 'liquid'
            else:
                static_max_any_psig[i] = float(gas_max)
                static_max_any_mp[i] = float(gas_max_mp)
                static_max_any_elev_ft[i] = float(gas_max_elev)
                static_max_any_phase[i] = 'gas'
        
            if (maop_psig is not None) and np.isfinite(static_max_any_psig[i]) and (float(static_max_any_psig[i]) > float(maop_psig) + 1e-6):
                static_violations.append({
                    'time_hr': float(elapsed_times[i]),
                    'pig_mp': float(pig_mp_now),
                    'phase': str(static_max_any_phase[i]),
                    'max_static_psig': float(static_max_any_psig[i]),
                    'max_static_mp': float(static_max_any_mp[i]),
                    'max_static_elev_ft': float(static_max_any_elev_ft[i]),
                    'maop_psig': float(maop_psig),
                    'margin_psig': float(static_max_any_psig[i]) - float(maop_psig),
                })

        # --- Evaluate monitor point pressures at this step (piecewise if IPS check is sealed) ---
        for p in monitor_points:
            mp_p = float(p['mp'])
            elev_p = float(p.get('elev', 0.0))
            label_p = p['name']

            # Determine whether this monitor location is in the liquid-filled section (ahead of the pig)
            is_liquid = (mp_p >= pig_mp_now - 1e-9)

            if is_liquid:
                if use_ips and (active_ips is not None) and (mp_p >= float(active_ips) - 1e-9):
                    # Downstream of the sealed IPS: reference to endpoint boundary
                    mp_ref = float(cfg['system_end_mp'])
                    elev_ref = float(h_exit)
                    P_ref = float(P_end)
                elif use_ips and (active_ips is not None):
                    # Upstream of the sealed IPS (but ahead of pig): reference to IPS suction
                    mp_ref = float(active_ips)
                    elev_ref = float(elev_boundary_use)
                    P_ref = float(P_boundary_use)
                else:
                    # No IPS boundary: reference to endpoint
                    mp_ref = float(cfg['system_end_mp'])
                    elev_ref = float(h_exit)
                    P_ref = float(P_end)

                L_seg_ft = max(0.0, (mp_ref - mp_p) * 5280.0)
                fr_seg = slug_friction_psi_SI(L_seg_ft, pipe_diameter, v, api_for_dp, viscosity_cst, roughness, sg_fallback=specific_gravity) if v > 1e-12 else 0.0
                head_seg = float(fluid_density) * (elev_ref - elev_p) / 144.0
                P_p = float(P_ref) + head_seg + fr_seg
            else:
                # Gas-filled section: approximate with behind-pig pressure (conservative for MAOP at upstream gas side)
                P_p = float(P_behind_pig_out)

            monitor_pressures[p['name']][i] = P_p


            # Static (blocked-in) hypothetical pressure at this monitor point
            if np.isfinite(hgl_static_liq_ft):
                if mp_p >= pig_mp_now - 1e-9:
                    monitor_pressures_static[p['name']][i] = (hgl_static_liq_ft - elev_p) * float(fluid_density) / 144.0
                else:
                    monitor_pressures_static[p['name']][i] = float(gas_static_psig)
            # Monitor rule checks (only apply slack rule to liquid-filled section)
            max_lim = p.get('max_psig', None)
            if max_lim is not None and P_p > max_lim + 1e-6:
                exceed_points.append({
                    'mp': mp_p,
                    'label': label_p,
                    'rule': 'MAOP',
                    'limit_psig': max_lim,
                    'pressure_psig': P_p,
                    'margin_psig': P_p - max_lim
                })
                monitor_violations.append({
                    'time_hr': float(elapsed_times[i]),
                    'pig_mp': pig_mp_now,
                    'point_name': p['name'],
                    'point_mp': mp_p,
                    'point_elev': elev_p,
                    'phase': 'liquid' if is_liquid else 'gas',
                    'rule': 'MAX_PRESSURE',
                    'limit_psig': float(max_lim),
                    'pressure_psig': float(P_p),
                    'margin_psig': float(P_p) - float(max_lim)
                })
            min_lim = p.get('min_psig', None)
            if (min_lim is not None) and is_liquid and (P_p < min_lim - 1e-6):
                exceed_points.append({
                    'mp': mp_p,
                    'label': label_p,
                    'rule': 'MIN',
                    'limit_psig': min_lim,
                    'pressure_psig': P_p,
                    'margin_psig': P_p - min_lim
                })
                monitor_violations.append({
                    'time_hr': float(elapsed_times[i]),
                    'pig_mp': pig_mp_now,
                    'point_name': p['name'],
                    'point_mp': mp_p,
                    'point_elev': elev_p,
                    'phase': 'liquid' if is_liquid else 'gas',
                    'rule': 'MIN_PRESSURE',
                    'limit_psig': float(min_lim),
                    'pressure_psig': float(P_p),
                    'margin_psig': float(P_p) - float(min_lim)
                })

            # Global slack check (liquid-filled section only)
            if is_liquid and (P_p < float(slack_threshold_psig) - 1e-6):
                exceed_points.append({
                    'mp': mp_p,
                    'label': label_p,
                    'rule': 'SLACK',
                    'limit_psig': float(slack_threshold_psig),
                    'pressure_psig': P_p,
                    'margin_psig': P_p - float(slack_threshold_psig)
                })
                monitor_violations.append({
                    'time_hr': float(elapsed_times[i]),
                    'pig_mp': pig_mp_now,
                    'point_name': p['name'],
                    'point_mp': mp_p,
                    'point_elev': elev_p,
                    'phase': 'liquid',
                    'rule': 'SLACK',
                    'limit_psig': float(slack_threshold_psig),
                    'pressure_psig': float(P_p),
                    'margin_psig': float(P_p) - float(slack_threshold_psig)
                })

        # Time for this segment; if v too small, stop
        if v <= 1e-8:
            last_valid_i = i
            logging.warning(f"Near-zero velocity at MP {purge_mileposts[i]:.2f}; stopping.")
            break
        time_step_hours = segment_length_ft / v / 3600.0
        elapsed_times[i + 1] = elapsed_times[i] + time_step_hours

        if elapsed_times[i + 1] > max_elapsed_time_hr:
            stop_reason = f"MAX_ELAPSED_TIME>{max_elapsed_time_hr:.0f}hr"
            last_valid_i = min(i + 1, n_points - 1)
            logging.warning(
                f"Max elapsed time cap reached ({max_elapsed_time_hr:.0f} hr) at MP={purge_mileposts[last_valid_i]:.3f}. "
                "Stopping early."
            )
            break

        # Integrate nitrogen (SCF @ std) during injection
        if coast_mode and cutoff_value is not None:
            cumulative_n2[i] = cutoff_value
        n2_added_scf = 0.0 if coast_mode else (q_use * time_step_hours * 60.0)
        cumulative_n2[i + 1] = cumulative_n2[i] + n2_added_scf

        # Update trapped gas volume (coast) after moving
        if coast_mode and V_gas_ft3 is not None:
            V_gas_ft3 += area * segment_length_ft

        cumulative_distance += segment_length_ft
        # Switch to coasting for next steps once cutoff reached
        if (not coast_mode) and (cutoff_value is not None) and (cumulative_n2[i + 1] >= cutoff_value):
            cumulative_n2[i + 1] = cutoff_value
            coast_mode = True
            n2_moles_total = n2_moles_from_scf(cutoff_value)
            # Initialize gas volume behind pig at the *next* position
            L_gas_next_ft = max(1e-6, (purge_mileposts[i + 1] - cfg['purge_start_mp']) * 5280.0)
            V_gas_ft3 = max(1e-6, area * L_gas_next_ft)



    # If we exited because of an explicit step cap (not a stall/timeout), mark it.
    if stop_reason is None and max_steps < (n_points - 1):
        stop_reason = f"MAX_STEPS({max_steps})"
        last_valid_i = min(max_steps, n_points - 1)
        logging.warning(f"Simulation stopped due to max_sim_steps cap: {stop_reason}")
    # --- Populate the final row (pig at purge end) so exports/tables don't show zeros ---
    j = min(last_valid_i, n_points - 1)
    try:
        mp_j = purge_mileposts[j]
        P_exit_j = target_exit_pressure(mp_j, cfg, cfg['purge_start_mp'], cfg['purge_end_mp'], cfg['throttle_down_miles'])
        head_j = head_losses[j]
        # With pig stopped, flowing friction is ~0; static requirement is exit + head (if uphill)
        P_behind_pig_static = max(P_exit_j, P_exit_j + head_j)

        if coast_mode and (n2_moles_total is not None) and (V_gas_ft3 is not None):
            p_psia = n2_pressure_psia_from_moles(n2_moles_total, V_gas_ft3, n2_temp_f)
            P_trap_psig = max(0.0, p_psia - ATM_PSI)
            behind_pig_pressures[j] = P_trap_psig
            injection_pressures[j] = P_trap_psig
            gas_dp_losses[j] = 0.0
        else:
            behind_pig_pressures[j] = P_behind_pig_static
            injection_pressures[j] = P_behind_pig_static
            gas_dp_losses[j] = 0.0

        exit_pressures[j] = P_exit_j
        friction_losses[j] = 0.0
        pig_speeds[j] = 0.0
        injection_rates[j] = 0.0
        if coast_mode and cutoff_value is not None:
            cumulative_n2[j] = cutoff_value
    except Exception:
        pass    # --- Final monitor evaluation at the last populated index j (static) ---
    # Note: per-step monitor violations are already tracked in-loop. This final pass is only to
    # populate the last monitor pressure sample (n_points-1) which is not visited inside the main loop.
    try:
        if monitor_points:
            pig_mp_now = float(purge_mileposts[j])
            for p in monitor_points:
                mp_p = float(p['mp'])
                elev_p = float(p.get('elev', cs(mp_p)))
                phase = 'liquid' if (mp_p >= pig_mp_now - 1e-9) else 'gas'
                if phase == 'liquid':
                    head_p = fluid_density * (float(h_exit) - elev_p) / 144.0
                    P_p = float(exit_pressures[j]) + float(head_p)
                else:
                    P_p = float(behind_pig_pressures[j])
                monitor_pressures[p['name']][j] = float(P_p)
    except Exception:
        pass


    # --- Build alarms summary (simple operational flags) ---
    alarms = []
    alarm_events = []

    # Stall / stop condition (near-zero velocity)
    if last_valid_i < (n_points - 1):
        alarms.append('STALL')
        alarm_events.append({
            'alarm': 'STALL',
            'time_hr': float(elapsed_times[last_valid_i]),
            'pig_mp': float(purge_mileposts[last_valid_i]),
            'detail': 'Near-zero velocity; simulation stopped early.'
        })


    # If we stopped for another reason, include it in the alarm detail.
    if stop_reason and 'TIMEOUT' in str(stop_reason):
        alarms.append('TIMEOUT')
        alarm_events.append({
            'alarm': 'TIMEOUT',
            'time_hr': float(elapsed_times[last_valid_i]) if last_valid_i >= 0 else 0.0,
            'pig_mp': float(purge_mileposts[last_valid_i]) if last_valid_i >= 0 else float(purge_mileposts[0]),
            'detail': f"Stopped due to wall-clock timeout ({stop_reason})."
        })
    if stop_reason and 'MAX_ELAPSED_TIME' in str(stop_reason):
        alarms.append('MAX_ELAPSED_TIME')
        alarm_events.append({
            'alarm': 'MAX_ELAPSED_TIME',
            'time_hr': float(elapsed_times[last_valid_i]) if last_valid_i >= 0 else 0.0,
            'pig_mp': float(purge_mileposts[last_valid_i]) if last_valid_i >= 0 else float(purge_mileposts[0]),
            'detail': f"Stopped due to max elapsed-time cap ({stop_reason})."
        })
    if stop_reason and 'MAX_STEPS' in str(stop_reason):
        alarms.append('MAX_STEPS')
        alarm_events.append({
            'alarm': 'MAX_STEPS',
            'time_hr': float(elapsed_times[last_valid_i]) if last_valid_i >= 0 else 0.0,
            'pig_mp': float(purge_mileposts[last_valid_i]) if last_valid_i >= 0 else float(purge_mileposts[0]),
            'detail': f"Stopped due to explicit step cap ({stop_reason})."
        })

    # Minimum pig speed alarm (not enforced; alarm only)
    try:
        min_speed_alarm_mph = float(cfg.get('min_pig_speed')) if cfg.get('min_pig_speed') is not None else None
    except Exception:
        min_speed_alarm_mph = None

    if min_speed_alarm_mph is not None and last_valid_i >= 0:
        speeds = np.asarray(pig_speeds[:last_valid_i + 1], dtype=float)
        if np.any(np.isfinite(speeds)):
            jmin = int(np.nanargmin(speeds))
            vmin = float(speeds[jmin])
            if vmin < (min_speed_alarm_mph - 1e-9):
                alarms.append('BELOW_MIN_SPEED')
                alarm_events.append({
                    'alarm': 'BELOW_MIN_SPEED',
                    'time_hr': float(elapsed_times[jmin]),
                    'pig_mp': float(purge_mileposts[jmin]),
                    'detail': f"Pig speed fell below min_pig_speed alarm threshold ({min_speed_alarm_mph:.3g} mph).",
                    'value': vmin,
                    'limit': min_speed_alarm_mph
                })

    # Monitor-based alarms (slack / MAOP / min pressure at IPS)
    try:
        mv = monitor_violations or []
        if any(v.get('rule') == 'SLACK' for v in mv):
            alarms.append('SLACK')
        if any(v.get('rule') == 'MAX_PRESSURE' for v in mv):
            alarms.append('MAOP_EXCEED')
        if any(v.get('rule') == 'MIN_PRESSURE' for v in mv):
            alarms.append('MIN_PRESSURE')
    except Exception:
        pass

    # de-dup while preserving order
    _seenA = set()
    alarms = [a for a in alarms if not (a in _seenA or _seenA.add(a))]
    logging.info("Exiting run_simulation")
    return {
        'purge_mileposts': purge_mileposts[:last_valid_i + 1],
        'elevations': elevations[:last_valid_i + 1],
        'elapsed_times': elapsed_times[:last_valid_i + 1],
        # Preferred naming (used throughout UI/output):
        'behind_pig_pressures': behind_pig_pressures[:last_valid_i + 1],  # psig immediately behind pig
        # Backward-compatible alias (internal legacy name):
        'drive_pressures': behind_pig_pressures[:last_valid_i + 1],
        'injection_pressures': injection_pressures[:last_valid_i + 1],
        'gas_dp_losses': gas_dp_losses[:last_valid_i + 1],
        'friction_losses': friction_losses[:last_valid_i + 1],
        'head_losses': head_losses[:last_valid_i + 1],
        'exit_pressures': exit_pressures[:last_valid_i + 1],
        'exit_pressures_programmed': exit_pressures_programmed[:last_valid_i + 1],
        'exit_pressures_lb': exit_pressures_lb[:last_valid_i + 1],
        'exit_pressures_ub': exit_pressures_ub[:last_valid_i + 1],
        'injection_rates': injection_rates[:last_valid_i + 1],
        'cumulative_n2': cumulative_n2[:last_valid_i + 1],
        'differential_pressures': differential_pressures[:last_valid_i + 1],
        'miles_to_outlet': miles_to_outlet[:last_valid_i + 1],
        'pig_speeds': pig_speeds[:last_valid_i + 1],
        'monitor_points': monitor_points,
        'monitor_pressures': {k: v[:last_valid_i + 1] for k, v in monitor_pressures.items()},
        'monitor_violations': monitor_violations,
        'monitor_pressures_static': monitor_pressures_static,
        'static_hgl_ft': static_hgl_ft[:last_valid_i + 1],
        'static_max_any_psig': static_max_any_psig[:last_valid_i + 1],
        'static_max_any_mp': static_max_any_mp[:last_valid_i + 1],
        'static_max_any_elev_ft': static_max_any_elev_ft[:last_valid_i + 1],
        'static_max_any_phase': static_max_any_phase[:last_valid_i + 1],
        'static_violations': static_violations,
        'slack_threshold_psig': slack_threshold_psig,
        'maop_psig': maop_psig,
        'exceed_points': exceed_points,
        'hydraulic_boundary_mps': hydraulic_boundary_mps[:last_valid_i + 1],
        'hydraulic_boundary_pressures': hydraulic_boundary_pressures[:last_valid_i + 1],
        'head_to_boundary': head_to_boundary[:last_valid_i + 1],
        'ips_active_mp': ips_active_mp[:last_valid_i + 1],
        'ips_suction_psig': ips_suction_psig[:last_valid_i + 1],
        'ips_discharge_req_psig': ips_discharge_req_psig[:last_valid_i + 1],
        'ips_pump_dp_req_psi': ips_pump_dp_req_psi[:last_valid_i + 1],
        'ips_flow_bph': ips_flow_bph[:last_valid_i + 1],
        'ips_state': ips_state[:last_valid_i + 1],
        'alarms': alarms,
        'alarm_events': alarm_events,
        'last_valid_i': last_valid_i
    }

def visualize_results(results, inputs):
    logging.info("Entering visualize_results")
    fig, (ax1, ax2, ax3, ax4) = plt.subplots(4, 1, figsize=(10, 15))
    
    ax1.plot(results['purge_mileposts'] - inputs['purge_start_mp'], results['elevations'], 'b-')
    ax1.set_title('Elevation vs. Milepost')
    ax1.set_xlabel('Miles')
    ax1.set_ylabel('Elevation (ft)')
    
    ax2.plot(results['purge_mileposts'] - inputs['purge_start_mp'], results['pig_speeds'], 'r-')
    ax2.axhline(y=inputs['max_pig_speed'] * 1.25, linestyle='--', label='Max Speed × 1.25')
    ax2.axhline(y=inputs['target_pig_speed'], linestyle=':', label='Target Speed')
    # Optional: show equivalent speed cap from Max Outlet Flow constraint
    try:
        _cap = inputs.get('max_pig_speed_from_max_outlet_flow_mph', None)
        if _cap is not None and _cap != "N/A":
            _cap = float(_cap)
            ax2.axhline(y=_cap, linestyle='-.', label='Max Flow Speed Cap')
    except Exception:
        pass
    ax2.set_title('Pig Speed vs. Milepost')
    ax2.set_xlabel('Miles')
    ax2.set_ylabel('Speed (mph)')
    ax2.legend()
    
    ax3.plot(results['purge_mileposts'] - inputs['purge_start_mp'], results['injection_pressures'], label='Injection Pressure (inlet)')
    behind = results.get('behind_pig_pressures', results.get('drive_pressures'))
    ax3.plot(results['purge_mileposts'] - inputs['purge_start_mp'], behind, label='Pressure Behind Pig')
    ax3.plot(results['purge_mileposts'] - inputs['purge_start_mp'], results['exit_pressures'], label='Exit Pressure (target/taper)')
    ax3.plot(results['purge_mileposts'] - inputs['purge_start_mp'], results['gas_dp_losses'], label='Gas DP inlet→pig')
    ax3.set_title('Pressure vs. Milepost')
    ax3.set_xlabel('Miles')
    ax3.set_ylabel('Pressure (psi)')
    ax3.legend()
    
    ax4.plot(results['purge_mileposts'] - inputs['purge_start_mp'], results['cumulative_n2'], 'c-')
    ax4.set_title('Cumulative Nitrogen vs. Milepost')
    ax4.set_xlabel('Miles')
    ax4.set_ylabel('N2 Volume (SCF)')
    
    plt.tight_layout()
    plt.show()
    logging.info("Exiting visualize_results")

def print_condensed_table(results, inputs):
    logging.info("Printing condensed table")
    # pipe area for throughput
    nps_key = resolve_nps_key(inputs['nps'])
    pipe_od = nps_data[nps_key]["OD_in"]
    pipe_id = pipe_od - 2 * float(inputs['pipe_wt'])
    pipe_diameter_ft = pipe_id / 12.0
    area = math.pi * (pipe_diameter_ft / 2) ** 2
    print("\nCondensed Table (50 evenly spaced points):")
    print(
        "Miles\tElevation (ft)\tElapsed Time (hr)\tInjection P (psi)\tGas DP (psi)\tBehind Pig P (psi)\t"
        "Liquid Fric (psi)\tHead (psi)\tExit P (psi)\tInjection Rate (SCFM)\tCumulative N2 (SCF)\t"
        "Pig Speed (mph)\tMiles to Outlet\tBarrels/hr"
    )
    
    n_points = len(results['purge_mileposts'])
    indices = np.linspace(0, results['last_valid_i'], min(50, n_points), dtype=int)
    
    for i in indices:
        miles = results['purge_mileposts'][i] - inputs['purge_start_mp']
        elevation = results['elevations'][i]
        time_hr = results['elapsed_times'][i]
        behind = results.get('behind_pig_pressures', results.get('drive_pressures'))
        inj_p = results.get('injection_pressures', np.zeros_like(behind))[i]
        gas_dp = results.get('gas_dp_losses', np.zeros_like(behind))[i]
        behind_p = behind[i]
        friction = results['friction_losses'][i]
        head = results['head_losses'][i]
        exit_p = results['exit_pressures'][i]
        inj_rate = results['injection_rates'][i]
        n2_vol = results['cumulative_n2'][i]
        speed = results['pig_speeds'][i]
        miles_to_outlet = inputs['system_end_mp'] - results['purge_mileposts'][i]
        bbl_hr = speed * 5280.0 * area / 5.614583
        print(
            f"{miles:.2f}\t{elevation:.2f}\t{time_hr:.2f}\t\t"
            f"{inj_p:.0f}\t\t{gas_dp:.1f}\t\t{behind_p:.0f}\t\t"
            f"{friction:.2f}\t\t{head:.0f}\t\t{exit_p:.0f}\t\t"
            f"{int(inj_rate):,}\t\t{int(n2_vol):,}\t\t{speed:.2f}\t\t{miles_to_outlet:.2f}\t\t{bbl_hr:.0f}"
        )

def print_alarm_summary(results, inputs):
    """Print a concise alarm summary after a run."""
    alarms = results.get('alarms', []) or []
    mv = results.get('monitor_violations', []) or []
    alarm_events = results.get('alarm_events', []) or []

    print("\nALARM SUMMARY:")
    if not alarms:
        print("  No alarms tripped.")
        return

    # Helper to pull worst-case event for a given rule
    def _worst(rule, kind='min'):
        evs = [v for v in mv if v.get('rule') == rule]
        if not evs:
            return None
        # margin_psig is (pressure - limit); for MAX_PRESSURE exceed >0 is bad, for MIN/SLACK <0 is bad
        if kind == 'max':
            return max(evs, key=lambda x: float(x.get('margin_psig', float('-inf'))))
        return min(evs, key=lambda x: float(x.get('margin_psig', float('inf'))))

    # Stall / below min speed from alarm_events
    for a in alarms:
        if a == 'STALL':
            e = next((x for x in alarm_events if x.get('alarm') == 'STALL'), None)
            if e:
                print(f"  STALL: stopped early at MP {e.get('pig_mp'):.3f} (t={e.get('time_hr'):.2f} hr).")
            else:
                print("  STALL: stopped early (near-zero velocity).")

        elif a == 'BELOW_MIN_SPEED':
            e = next((x for x in alarm_events if x.get('alarm') == 'BELOW_MIN_SPEED'), None)
            if e:
                print(f"  MIN SPEED: {e.get('value'):.3g} mph < {e.get('limit'):.3g} mph at MP {e.get('pig_mp'):.3f} (t={e.get('time_hr'):.2f} hr).")
            else:
                print("  MIN SPEED: fell below min_pig_speed threshold at least once.")

        elif a == 'SLACK':
            w = _worst('SLACK', kind='min')
            if w:
                print(f"  SLACK: {w.get('pressure_psig'):.2f} psig < {w.get('limit_psig'):.2f} psig at {w.get('point_name')} (MP {w.get('point_mp'):.3f}) "
                      f"t={w.get('time_hr'):.2f} hr, margin {w.get('margin_psig'):.2f} psi.")
            else:
                print("  SLACK: slack threshold violated at least once.")

        elif a == 'MAOP_EXCEED':
            w = _worst('MAX_PRESSURE', kind='max')
            if w:
                print(f"  MAOP: {w.get('pressure_psig'):.2f} psig > {w.get('limit_psig'):.2f} psig at {w.get('point_name')} (MP {w.get('point_mp'):.3f}) "
                      f"t={w.get('time_hr'):.2f} hr, exceed {w.get('margin_psig'):.2f} psi.")
            else:
                print("  MAOP: MAOP exceeded at least once.")

        elif a == 'MIN_PRESSURE':
            w = _worst('MIN_PRESSURE', kind='min')
            if w:
                print(f"  MIN PRESSURE: {w.get('pressure_psig'):.2f} psig < {w.get('limit_psig'):.2f} psig at {w.get('point_name')} (MP {w.get('point_mp'):.3f}) "
                      f"t={w.get('time_hr'):.2f} hr, margin {w.get('margin_psig'):.2f} psi.")
            else:
                print("  MIN PRESSURE: minimum pressure constraint violated at least once.")


def export_results(dialog_root, inputs, results):
    logging.info("Entering export_results")
    # Resolve NPS key safely (fixes '16.0' KeyError)
    try:
        nps_key = resolve_nps_key(inputs['nps'])
    except Exception:
        # fall back to string
        nps_key = str(inputs['nps'])

    inputs_data = {
        "Parameter": [
            "Nominal Pipe Size (NPS)",
            "Outside Diameter (in)",
            "Wall Thickness (in)",
            "Pipe Material",
            "Fluid Type",
            "API Gravity (if Crude)",
            "Viscosity (cSt) (if Crude)",
            "Max Nitrogen Drive Pressure (psig)",
            "Max N2 Rate (SCFM)",
            "Min Exit Pressure (Run) (psig)",
            "Exit Pressure (End) (psig)",
            "Exit Pressure Behavior",
            "Endpoint Constraint Mode",
            "Exit Pressure Min Clamp (psig)",
            "Exit Pressure Max Clamp (psig)",
            "Exit Pressure Ramp Limit (psi/hr)",
            "Slack Threshold (psig)",
            "MAOP (psig)",
            "Max Outlet Flow to Tankage (BPH)",
            "Equivalent Max Pig Speed from Max Flow (mph)",
            "Max Pig Speed (mph)",
            "Min Pig Speed (mph)",
            "Target Pig Speed (mph)",
            "Purge Start Milepost",
            "Purge End Milepost",
            "System Endpoint Milepost",
            "Throttle Down Distance (miles)",
            "Estimated N2 Pressure at End (psig)",
            "Has Intermediate Pump Stations",
            "IPS Mileposts",
            "IPS Shutdown Distance (miles)",
            "Minimum Pump Suction Pressure (psig)",
            "Minimum Pump Flow (BPH)",
            "IPS Check Valve Seal ΔP (psi)",
            "Resolution (points)",
            "Resample purge profile to resolution"
        ],
        "Value": [
            inputs.get('nps', "N/A"),
            nps_data[nps_key]["OD_in"],
            inputs.get('pipe_wt', "N/A"),
            roughness_data[int(inputs['roughness_num'])]['material'] if inputs.get('roughness_num', None) is not None else "N/A",
            fluid_data[int(inputs['fluid_num'])]['name'] if inputs.get('fluid_num', None) is not None else "N/A",
            inputs.get('api_gravity', "N/A"),
            inputs.get('viscosity_cst', "N/A"),
            inputs.get('max_nitrogen_pressure', inputs.get('max_drive_pressure', "N/A")),
            inputs.get('max_n2_rate_scfm', "N/A"),
            inputs.get('exit_pressure_run', "N/A"),
            inputs.get('exit_pressure_end', "N/A"),
            inputs.get('exit_behavior', "N/A"),
            inputs.get('endpoint_constraint_mode', "none"),
            inputs.get('exit_pressure_min_clamp', "N/A"),
            inputs.get('exit_pressure_max_clamp', "N/A"),
            inputs.get('exit_pressure_ramp_psi_per_hr', "N/A"),
            inputs.get('slack_pressure_psig', 50.0),
            inputs.get('maop_psig', "N/A"),
            inputs.get('max_outlet_flow_bph', "N/A"),
            inputs.get('max_pig_speed_from_max_outlet_flow_mph', "N/A"),
            inputs.get('max_pig_speed', "N/A"),
            inputs.get('min_pig_speed', "N/A"),
            inputs.get('target_pig_speed', "N/A"),
            inputs.get('purge_start_mp', "N/A"),
            inputs.get('purge_end_mp', "N/A"),
            inputs.get('system_end_mp', "N/A"),
            inputs.get('throttle_down_miles', "N/A"),
            inputs.get('n2_end_pressure', "N/A"),
            inputs.get('has_ips', False),
            inputs.get('ips_mp', inputs.get('ips_mps', "N/A")) or "N/A",
            inputs.get('ips_shutdown_dist', "N/A"),
            inputs.get('min_pump_suction_pressure', "N/A"),
            inputs.get('min_pump_flow_bph', "N/A"),
            inputs.get('ips_check_dp_psi', "N/A"),
            inputs.get('resolution', "N/A"),
            inputs.get('resample_profile_to_resolution', True),
        ]
    }
    inputs_df = pd.DataFrame(inputs_data)
    
    # area for bbl/hr and miles to outlet
    nps_key2 = resolve_nps_key(inputs['nps'])
    pipe_od2 = nps_data[nps_key2]["OD_in"]
    pipe_id2 = pipe_od2 - 2 * float(inputs['pipe_wt'])
    pipe_diameter_ft2 = pipe_id2 / 12.0
    area2 = math.pi * (pipe_diameter_ft2 / 2) ** 2
    miles_to_outlet = inputs['system_end_mp'] - results['purge_mileposts']
    bbl_hr = results['pig_speeds'] * 5280.0 * area2 / 5.614583

    full_results_df = pd.DataFrame({
        "Miles": results['purge_mileposts'] - inputs['purge_start_mp'],
        "Elevation (ft)": results['elevations'],
        "Elapsed Time (hours)": results['elapsed_times'],
        "Injection Pressure (psi)": results.get('injection_pressures', results.get('behind_pig_pressures', results['drive_pressures'])),
        "Gas DP Inlet→Pig (psi)": results.get('gas_dp_losses', np.zeros_like(results.get('behind_pig_pressures', results['drive_pressures']))),
        "Pressure Behind Pig (psi)": results.get('behind_pig_pressures', results['drive_pressures']),
        "Friction Loss (psi)": results['friction_losses'],
        "Head Pressure (psi)": results['head_losses'],
        "Exit Pressure (psi)": results['exit_pressures'],
        "Injection Rate (SCFM)": results['injection_rates'],
        "Cumulative N2 (SCF)": results['cumulative_n2'],
        "Pig Speed (mph)": results['pig_speeds'],
        "Miles to Outlet": miles_to_outlet,
        "Barrels per hour": bbl_hr
    })
    
    n_points = len(results['purge_mileposts'])
    condensed_indices = np.linspace(0, results['last_valid_i'], min(50, n_points), dtype=int)
    condensed_results_df = full_results_df.iloc[condensed_indices]
    
    try:
        dialog_root.update()
        start_time = time.time()
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            parent=dialog_root
        )
        dialog_root.update()
        if time.time() - start_time > 45:
            logging.warning("Export file dialog timeout after 45 seconds")
            raise ValueError("Export file dialog timed out. Try running from Command Prompt: cd C:\\Users\\YourName\\Documents && python Purge_Modeling_Program_22_09_2025.py")
        logging.info(f"Selected output file: {file_path}")
        if file_path:
            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                inputs_df.to_excel(writer, sheet_name="Simulation Inputs", index=False)
                full_results_df.to_excel(writer, sheet_name="Full Results", index=False)
                condensed_results_df.to_excel(writer, sheet_name="Condensed Results", index=False)

                # Monitoring outputs (if enabled / available)
                monitor_points = results.get('monitor_points', []) or []
                monitor_pressures = results.get('monitor_pressures', {}) or {}
                monitor_pressures_static = results.get('monitor_pressures_static', {}) or {}
                static_violations = results.get('static_violations', []) or []
                monitor_violations = results.get('monitor_violations', []) or []

                if monitor_points and monitor_pressures:
                    # --- Monitors & time-series exports (dynamic + static) ---
                    ordered = sorted(monitor_points, key=lambda p: float(p.get('mp', 0.0)))

                    # Monitors (metadata)
                    monitor_points_df = pd.DataFrame([{
                        'Name': p.get('name'),
                        'Scope': p.get('scope', ''),
                        'Category': p.get('category'),
                        'Milepost': p.get('mp'),
                        'Elevation (ft)': p.get('elev'),
                        'Min Rule (psig)': p.get('min_psig'),
                        'Max Rule (psig)': p.get('max_psig'),
                    } for p in ordered])
                    monitor_points_df.to_excel(writer, sheet_name="Monitors", index=False)

                    # Shared base columns for series sheets
                    n_steps = len(results['purge_mileposts'])
                    base = pd.DataFrame({
                        'Elapsed Time (hr)': results['elapsed_times'],
                        'Pig Milepost': results['purge_mileposts'],
                        'Miles into purge': results['purge_mileposts'] - inputs['purge_start_mp'],
                        'Exit Pressure (psig)': results['exit_pressures'],
                        'Hydraulic Boundary MP': results.get('hydraulic_boundary_mp', np.full(n_steps, np.nan, dtype=float)),
                        'Miles to Hydraulic Boundary': results.get('hydraulic_boundary_mp', np.full(n_steps, np.nan, dtype=float)) - results['purge_mileposts'],
                        'Hydraulic Boundary Pressure (psig)': results.get('hydraulic_boundary_pressure', np.full(n_steps, np.nan, dtype=float)),
                        'Head to Boundary (psi)': results.get('head_to_boundary_psi', np.full(n_steps, np.nan, dtype=float)),
                        'IPS State': results.get('ips_state', np.array([''] * n_steps, dtype=object)),
                        'IPS Active MP': results.get('ips_active_mp', np.full(n_steps, np.nan, dtype=float)),
                        'IPS Suction (psig)': results.get('ips_suction_psig', np.full(n_steps, np.nan, dtype=float)),
                        'IPS Discharge Req (psig)': results.get('ips_discharge_req_psig', np.full(n_steps, np.nan, dtype=float)),
                        'IPS Pump ΔP Req (psi)': results.get('ips_pump_dp_req_psi', np.full(n_steps, np.nan, dtype=float)),
                        'Mainline Flow (BPH)': results.get('ips_mainline_flow_bph', np.full(n_steps, np.nan, dtype=float)),
                    })

                    # Dynamic series: one column per monitor point
                    dyn = base.copy()
                    for p in ordered:
                        name = p.get('name')
                        if name in monitor_pressures:
                            dyn[name] = monitor_pressures[name]
                    dyn.to_excel(writer, sheet_name="Monitor Dynamic", index=False)

                    # Static (block-in) series: one column per monitor point
                    if monitor_pressures_static:
                        stat = base.copy()
                        for p in ordered:
                            name = p.get('name')
                            if name in monitor_pressures_static:
                                stat[name] = monitor_pressures_static[name]
                        stat.to_excel(writer, sheet_name="Monitor Static", index=False)

                    # Violations / alarms
                    if monitor_violations:
                        pd.DataFrame(monitor_violations).to_excel(writer, sheet_name="Violations Dynamic", index=False)
                    if static_violations:
                        pd.DataFrame(static_violations).to_excel(writer, sheet_name="Violations Static", index=False)

                    alarm_events = results.get('alarm_events', []) or []
                    if alarm_events:
                        pd.DataFrame(alarm_events).to_excel(writer, sheet_name="Alarms", index=False)

                    # Summary per monitor point (dynamic + static)
                    try:
                        pig_mps = np.asarray(results['purge_mileposts'], dtype=float)
                        times = np.asarray(results['elapsed_times'], dtype=float)
                        summary_rows = []
                        maop = inputs.get('maop_psig', None)
                        maop = float(maop) if maop is not None else None

                        for p in ordered:
                            name = p.get('name')
                            mp_p = float(p.get('mp'))
                            P_dyn = np.asarray(monitor_pressures.get(name, np.full_like(pig_mps, np.nan)), dtype=float)
                            P_sta = np.asarray(monitor_pressures_static.get(name, np.full_like(pig_mps, np.nan)), dtype=float)

                            if P_dyn.size != pig_mps.size:
                                continue

                            max_dyn = float(np.nanmax(P_dyn)) if np.any(np.isfinite(P_dyn)) else np.nan
                            max_sta = float(np.nanmax(P_sta)) if np.any(np.isfinite(P_sta)) else np.nan

                            # "Liquid-filled at the point" mask: point is ahead of pig (same convention used in the model)
                            mask_liq = mp_p >= pig_mps - 1e-9

                            min_dyn_liq = float(np.nanmin(P_dyn[mask_liq])) if np.any(mask_liq) and np.any(np.isfinite(P_dyn[mask_liq])) else np.nan
                            max_dyn_liq = float(np.nanmax(P_dyn[mask_liq])) if np.any(mask_liq) and np.any(np.isfinite(P_dyn[mask_liq])) else np.nan
                            max_sta_liq = float(np.nanmax(P_sta[mask_liq])) if np.any(mask_liq) and np.any(np.isfinite(P_sta[mask_liq])) else np.nan

                            min_lim = p.get('min_psig')
                            max_lim = p.get('max_psig')

                            # Worst margins (dynamic) w.r.t. min rule and max rule/MAOP
                            worst_min_margin = np.nan
                            worst_min_time = np.nan
                            worst_max_margin = np.nan
                            worst_max_time = np.nan

                            if min_lim is not None and np.any(mask_liq):
                                margins = P_dyn[mask_liq] - float(min_lim)
                                if np.any(np.isfinite(margins)):
                                    j = int(np.nanargmin(margins))
                                    worst_min_margin = float(margins[j])
                                    worst_min_time = float(times[mask_liq][j]) if times.size == pig_mps.size else np.nan

                            if max_lim is not None:
                                margins = float(max_lim) - P_dyn
                                if np.any(np.isfinite(margins)):
                                    j = int(np.nanargmin(margins))
                                    worst_max_margin = float(margins[j])
                                    worst_max_time = float(times[j]) if times.size == pig_mps.size else np.nan
                            elif maop is not None:
                                margins = maop - P_dyn
                                if np.any(np.isfinite(margins)):
                                    j = int(np.nanargmin(margins))
                                    worst_max_margin = float(margins[j])
                                    worst_max_time = float(times[j]) if times.size == pig_mps.size else np.nan

                            # Static worst margin to MAOP (block-in anywhere)
                            worst_sta_margin = np.nan
                            worst_sta_time = np.nan
                            if maop is not None and np.any(np.isfinite(P_sta)):
                                margins = maop - P_sta
                                j = int(np.nanargmin(margins))
                                worst_sta_margin = float(margins[j])
                                worst_sta_time = float(times[j]) if times.size == pig_mps.size else np.nan

                            summary_rows.append({
                                'Name': name,
                                'Milepost': mp_p,
                                'Elevation (ft)': float(p.get('elev', np.nan)),
                                'Category': p.get('category'),
                                'Max Dynamic (psig)': max_dyn,
                                'Min Dynamic (liquid) (psig)': min_dyn_liq,
                                'Max Dynamic (liquid) (psig)': max_dyn_liq,
                                'Max Static (block-in) (psig)': max_sta,
                                'Max Static (liquid) (psig)': max_sta_liq,
                                'Min Rule (psig)': min_lim,
                                'Worst MIN margin (dyn) (psig)': worst_min_margin,
                                'Worst MIN time (dyn) (hr)': worst_min_time,
                                'Max Rule (psig)': max_lim,
                                'Worst MAX margin (dyn) (psig)': worst_max_margin,
                                'Worst MAX time (dyn) (hr)': worst_max_time,
                                'MAOP (psig)': maop,
                                'Worst MAOP margin (static) (psig)': worst_sta_margin,
                                'Worst MAOP time (static) (hr)': worst_sta_time,
                            })

                        if summary_rows:
                            pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Monitor Summary", index=False)
                    except Exception:
                        pass

                    # Simple formatting (column widths, frozen header, filters)
                    try:
                        from openpyxl.styles import Font, Alignment
                        from openpyxl.utils import get_column_letter

                        def _fmt(ws):
                            ws.freeze_panes = "A2"
                            ws.auto_filter.ref = ws.dimensions
                            # header style
                            for cell in ws[1]:
                                cell.font = Font(bold=True)
                                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                            # column widths
                            for col in range(1, ws.max_column + 1):
                                max_len = 0
                                col_letter = get_column_letter(col)
                                for row in range(1, min(ws.max_row, 3000) + 1):
                                    v = ws.cell(row=row, column=col).value
                                    if v is None:
                                        continue
                                    s = str(v)
                                    if len(s) > max_len:
                                        max_len = len(s)
                                ws.column_dimensions[col_letter].width = max(10, min(48, max_len + 2))

                        for sh in ("Monitors", "Monitor Dynamic", "Monitor Static", "Violations Dynamic", "Violations Static", "Alarms", "Monitor Summary"):
                            if sh in writer.sheets:
                                _fmt(writer.sheets[sh])
                    except Exception:
                        pass

            messagebox.showinfo("Success", f"Results exported to {file_path}", parent=dialog_root)
        else:
            logging.warning("No output file selected")
            messagebox.showwarning("File Selection", "No output file selected. Results not saved.", parent=dialog_root)
    except Exception as e:
        logging.error(f"Failed to export results: {str(e)}")
        raise ValueError(f"Error exporting results: {str(e)}")
    logging.info("Exiting export_results")

def configure_profile_segments(dialog_root, file_paths):
    """Let the user review selected profile segments, choose which ones to invert, and set append order.

    Returns:
        List[Dict[str, Any]] like: [{'path': 'C:/...', 'invert': True}, ...]
    """
    file_paths = list(file_paths or [])
    if not file_paths:
        return []

    # Single file: just ask about inversion.
    if len(file_paths) == 1:
        fp = file_paths[0]
        invert = messagebox.askyesno(
            "Invert Profile Segment?",
            "Invert (reverse) this profile segment?\n\n"
            f"{os.path.basename(fp)}",
            parent=dialog_root
        )
        return [{"path": fp, "invert": bool(invert)}]

    # Multi-file: provide an order + invert editor.
    top = tk.Toplevel(dialog_root)
    top.title("Profile Segments: Order & Inversion")

    # Center on screen and force visibility (Windows can sometimes open dialogs behind other windows).
    w, h = 980, 420
    try:
        sw = top.winfo_screenwidth()
        sh = top.winfo_screenheight()
        x = max(50, (sw - w) // 2)
        y = max(50, (sh - h) // 3)
        top.geometry(f"{w}x{h}+{x}+{y}")
    except Exception:
        top.geometry("980x420")
    top.minsize(740, 320)

    # Force it to the front and grab focus
    try:
        top.update_idletasks()
        top.deiconify()
        top.lift()
        top.attributes('-topmost', True)
        top.focus_force()
        top.after(750, lambda: top.attributes('-topmost', False))
    except Exception:
        pass

    top.grab_set()
    logging.info("Profile segment order/inversion dialog opened; awaiting user action.")

    instructions = (
        "1) Review the segments you selected\n"
        "2) Toggle 'Invert' for any segment that runs backwards\n"
        "3) Use Up/Down to set the order they should be appended\n"
        "4) Click OK to continue"
    )
    ttk.Label(top, text=instructions, justify="left").pack(anchor="w", padx=12, pady=(12, 6))

    frame = ttk.Frame(top)
    frame.pack(fill="both", expand=True, padx=12, pady=6)

    columns = ("invert", "path")
    tree = ttk.Treeview(frame, columns=columns, show="headings", selectmode="extended")
    tree.heading("invert", text="Invert?")
    tree.heading("path", text="File")
    tree.column("invert", width=80, anchor="center", stretch=False)
    tree.column("path", width=820, anchor="w", stretch=True)

    yscroll = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
    xscroll = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
    tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)

    tree.grid(row=0, column=0, sticky="nsew")
    yscroll.grid(row=0, column=1, sticky="ns")
    xscroll.grid(row=1, column=0, sticky="ew")

    frame.rowconfigure(0, weight=1)
    frame.columnconfigure(0, weight=1)

    # Populate rows (default: not inverted)
    for fp in file_paths:
        tree.insert("", "end", values=("No", fp))

    btns = ttk.Frame(top)
    btns.pack(fill="x", padx=12, pady=(6, 12))

    def _selected_sorted():
        items = list(tree.selection())
        items.sort(key=lambda iid: tree.index(iid))
        return items

    def move(delta):
        items = _selected_sorted()
        if not items:
            return
        all_items = list(tree.get_children())
        n = len(all_items)

        if delta < 0:
            for iid in items:
                idx = tree.index(iid)
                new_idx = max(idx + delta, 0)
                tree.move(iid, "", new_idx)
        else:
            for iid in reversed(items):
                idx = tree.index(iid)
                new_idx = min(idx + delta, n - 1)
                tree.move(iid, "", new_idx)

    def toggle_invert():
        items = list(tree.selection())
        if not items:
            return
        for iid in items:
            cur = str(tree.set(iid, "invert")).strip().lower()
            new_val = "Yes" if cur in ("no", "false", "0", "") else "No"
            tree.set(iid, "invert", new_val)

    result = {"segments": None}

    def on_ok():
        segs = []
        for iid in tree.get_children():
            inv = str(tree.set(iid, "invert")).strip().lower() in ("yes", "true", "1", "y")
            fp = str(tree.set(iid, "path"))
            segs.append({"path": fp, "invert": inv})
        result["segments"] = segs
        top.destroy()

    def on_cancel():
        result["segments"] = None
        top.destroy()

    ttk.Button(btns, text="Move Up", command=lambda: move(-1)).pack(side="left", padx=(0, 6))
    ttk.Button(btns, text="Move Down", command=lambda: move(1)).pack(side="left", padx=(0, 12))
    ttk.Button(btns, text="Toggle Invert", command=toggle_invert).pack(side="left")

    ttk.Button(btns, text="Cancel", command=on_cancel).pack(side="right", padx=(6, 0))
    ttk.Button(btns, text="OK", command=on_ok).pack(side="right")

    top.protocol("WM_DELETE_WINDOW", on_cancel)
    top.bind("<Escape>", lambda e: on_cancel())
    top.wait_window()
    logging.info("Profile segment dialog closed.")

    if result["segments"] is None:
        raise ValueError("Profile segment selection cancelled.")
    return result["segments"]


def preview_profile_dialog(parent, mileposts, elevations, segment_bounds=None, elevation_units_label="ft"):
    """Modal preview of the combined elevation profile.

    Returns one of: 'approve', 'reselect', 'cancel'.
    """
    # Downsample very large profiles for snappy plotting (keeps shape, avoids UI lockups)
    mp = np.asarray(mileposts)
    el = np.asarray(elevations)
    if len(mp) > 30000:
        idx = np.linspace(0, len(mp) - 1, 30000, dtype=int)
        mp_plot = mp[idx]
        el_plot = el[idx]
    else:
        mp_plot = mp
        el_plot = el
    # Fallback if the TkAgg canvas isn't available for some reason:
    if FigureCanvasTkAgg is None or Figure is None:
        try:
            plt.figure(figsize=(10, 4))
            plt.plot(mp_plot, el_plot)
            plt.xlabel("Milepost (mi)")
            plt.ylabel(f"Elevation ({elevation_units_label})")
            plt.title("Pipeline Profile Preview")
            plt.grid(True, alpha=0.3)
            plt.tight_layout()
            plt.show()
        except Exception:
            pass

        resp = messagebox.askyesno(
            "Approve Profile?",
            "Approve this profile and continue?",
            parent=parent
        )
        return "approve" if resp else "reselect"

    top = tk.Toplevel(parent)
    top.title("Pipeline Profile Preview")
    top.geometry("1050x700")
    top.transient(parent)
    top.grab_set()

    # Header / stats
    hdr = ttk.Frame(top, padding=(12, 10, 12, 6))
    hdr.pack(fill="x")

    try:
        mp0 = float(np.min(mp))
        mp1 = float(np.max(mp))
        el0 = float(np.min(el))
        el1 = float(np.max(el))
        stats = f"MP: {mp0:.2f} → {mp1:.2f} mi   |   Elev: {el0:.0f} → {el1:.0f} {elevation_units_label}   |   Points: {len(mp)}"
    except Exception:
        stats = f"Points: {len(mp)}"

    ttk.Label(hdr, text=stats, font=("Segoe UI", 10, "bold")).pack(anchor="w")
    ttk.Label(
        hdr,
        text="Inspect the combined profile below. If something looks wrong, choose Re-select Segments to change invert/order before continuing.",
        wraplength=1000
    ).pack(anchor="w", pady=(4, 0))

    # Plot
    plot_frame = ttk.Frame(top, padding=(12, 6, 12, 6))
    plot_frame.pack(fill="both", expand=True)

    fig = Figure(figsize=(10, 5), dpi=100)
    ax = fig.add_subplot(111)
    ax.plot(mp_plot, el_plot)
    ax.set_xlabel("Milepost (mi)")
    ax.set_ylabel(f"Elevation ({elevation_units_label})")
    ax.grid(True, alpha=0.3)

    # Optional segment boundary markers
    if segment_bounds:
        try:
            y_min = float(np.min(el))
            y_max = float(np.max(el))
            y_span = (y_max - y_min) if (y_max > y_min) else 1.0
            y_label = y_max + 0.02 * y_span
        except Exception:
            y_label = None

        for b in segment_bounds:
            try:
                x0 = float(b["start_mp"])
                x1 = float(b["end_mp"])
                name = str(b.get("name", "segment"))
                inv = " (inv)" if b.get("invert") else ""
                ax.axvline(x0, linestyle="--", alpha=0.22)
                ax.axvline(x1, linestyle="--", alpha=0.22)
                if y_label is not None:
                    ax.text((x0 + x1) / 2.0, y_label, name + inv, ha="center", va="bottom", fontsize=8, alpha=0.85)
            except Exception:
                continue

    canvas = FigureCanvasTkAgg(fig, master=plot_frame)
    canvas.draw()
    canvas.get_tk_widget().pack(fill="both", expand=True)

    try:
        toolbar = NavigationToolbar2Tk(canvas, plot_frame)
        toolbar.update()
    except Exception:
        toolbar = None

    # Buttons
    btns = ttk.Frame(top, padding=(12, 8, 12, 12))
    btns.pack(fill="x")

    result = {"action": "cancel"}

    def on_approve():
        result["action"] = "approve"
        top.destroy()

    def on_reselect():
        result["action"] = "reselect"
        top.destroy()

    def on_cancel():
        result["action"] = "cancel"
        top.destroy()

    ttk.Button(btns, text="Re-select Segments", command=on_reselect).pack(side="left")
    ttk.Button(btns, text="Cancel", command=on_cancel).pack(side="right")
    ttk.Button(btns, text="Approve Profile", command=on_approve).pack(side="right", padx=(0, 8))

    top.protocol("WM_DELETE_WINDOW", on_cancel)
    top.wait_window()
    return result["action"]

def process_elevation_profile(dialog_root, initial_inputs):
    logging.info("Processing elevation profile")
    elevation_units = initial_inputs['elevation_units']
    
    try:
        dialog_root.update()
        start_time = time.time()
        files = filedialog.askopenfilenames(
            title="Select Elevation Profile Files (Multiple OK)",
            filetypes=[
                ("All Supported", ("*.kmz", "*.kml", "*.xlsx", "*.xls", "*.txt")),
                ("KMZ/KML", ("*.kmz", "*.kml")),
                ("Excel", ("*.xlsx", "*.xls")),
                ("TXT", ("*.txt",)),
            ],
            parent=dialog_root
        )
        dialog_root.update()
        logging.info(f"Selected files: {files}")
    except Exception as e:
        logging.error(f"Failed to open file dialog: {str(e)}")
        messagebox.showerror("File Dialog Error", f"Failed to open file dialog: {str(e)}. Try copying files to C:\\Users\\YourName\\Documents.", parent=dialog_root)
        raise ValueError(f"Failed to open file dialog: {str(e)}")
    
    if not files:
        logging.warning("No elevation profile files selected")
        messagebox.showerror("File Selection Error", "No elevation profile files selected.", parent=dialog_root)
        raise ValueError("No elevation profile files selected.")

    # Let the user choose inversion and append order for the selected segments
    segments = configure_profile_segments(dialog_root, files)

    
    mileposts = []
    elevations = []
    current_mp = 0.0
    segment_bounds = []
    
    for seg in segments:
        file_path = seg["path"]
        invert_segment = bool(seg.get("invert", False))
        seg_start_mp = current_mp
        logging.info(f"Processing file: {file_path} (invert={invert_segment})")
        if file_path.lower().endswith(('.kmz', '.kml')):
            try:
                parser = _get_secure_lxml_parser()
                if file_path.lower().endswith('.kmz'):
                    with zipfile.ZipFile(file_path, 'r') as z:
                        kml_files = [f for f in z.namelist() if f.lower().endswith('.kml')]
                        if not kml_files:
                            raise ValueError(f"No KML found in KMZ: {file_path}")
                        # Prefer doc.kml if present, otherwise first KML in sorted order
                        kml_choice = None
                        for f in kml_files:
                            if os.path.basename(f).lower() == 'doc.kml':
                                kml_choice = f
                                break
                        if kml_choice is None:
                            kml_choice = sorted(kml_files)[0]
                        with z.open(kml_choice) as kml_file:
                            tree = ET.parse(kml_file, parser=parser)
                else:
                    tree = ET.parse(file_path, parser=parser)
                
                root = tree.getroot()
                
                coords = extract_pipeline_coords_from_kml_root(
                    root, file_path=file_path, multiline_mode='longest', max_points=200000
                )
                if not coords:
                    raise ValueError(f"No valid coordinates found in {file_path}")
                
                if invert_segment:
                    coords = list(reversed(coords))

                
                file_mileposts = [current_mp]
                file_elevations = [coords[0][2]]
                for j in range(1, len(coords)):
                    try:
                        dist_miles = geodesic(coords[j-1][:2], coords[j][:2]).miles
                    except Exception as e:
                        logging.warning(f"Geodesic failed at point {j}: {str(e)}, using Haversine")
                        dist_miles = haversine_distance(coords[j-1][0], coords[j-1][1], coords[j][0], coords[j][1])
                    current_mp += dist_miles
                    file_mileposts.append(current_mp)
                    file_elevations.append(coords[j][2])
                
                mileposts.extend(file_mileposts)
                elevations.extend(file_elevations)
                segment_bounds.append({"name": os.path.basename(file_path), "start_mp": seg_start_mp, "end_mp": current_mp, "invert": invert_segment})
            except Exception as e:
                logging.error(f"Failed to parse KMZ/KML {file_path}: {str(e)}")
                raise ValueError(f"Failed to parse KMZ/KML {file_path}: {str(e)}")
        
        elif file_path.lower().endswith(('.xlsx', '.xls')):
            try:
                df = pd.read_excel(file_path)
                if 'Milepost' not in df.columns or 'Elevation' not in df.columns:
                    raise ValueError(f"Excel {file_path} must have 'Milepost' and 'Elevation' columns.")

                mp = pd.to_numeric(df['Milepost'], errors='coerce').to_numpy()
                el = pd.to_numeric(df['Elevation'], errors='coerce').to_numpy()
                valid = (~np.isnan(mp)) & (~np.isnan(el))
                mp = mp[valid]
                el = el[valid]
                if len(mp) < 2:
                    raise ValueError(f"Excel {file_path} must contain at least 2 valid points.")

                order = np.argsort(mp)
                mp = mp[order]
                el = el[order]

                # Rebase to 0 so multiple segments append cleanly.
                mp = mp - mp[0]

                if invert_segment:
                    mp_end = mp[-1]
                    mp_rev = mp[::-1]
                    el = el[::-1]
                    mp = mp_end - mp_rev

                file_mileposts = (mp + current_mp).tolist()
                file_elevations = el.tolist()
                current_mp = file_mileposts[-1]
                mileposts.extend(file_mileposts)
                elevations.extend(file_elevations)
                segment_bounds.append({"name": os.path.basename(file_path), "start_mp": seg_start_mp, "end_mp": current_mp, "invert": invert_segment})
            except Exception as e:
                logging.error(f"Failed to parse Excel {file_path}: {str(e)}")
                raise ValueError(f"Failed to parse Excel {file_path}: {str(e)}")
        
        elif file_path.lower().endswith('.txt'):
            try:
                df = pd.read_csv(file_path, sep=None, engine='python')
                if 'latitude' not in df.columns or 'longitude' not in df.columns or 'altitude (ft)' not in df.columns:
                    raise ValueError(f"TXT {file_path} must have 'latitude', 'longitude', and 'altitude (ft)' columns.")
                coords = list(zip(df['latitude'], df['longitude'], df['altitude (ft)']))
                if invert_segment:
                    coords = list(reversed(coords))

                file_mileposts = [current_mp]
                file_elevations = [coords[0][2]]
                for j in range(1, len(coords)):
                    try:
                        dist_miles = geodesic((coords[j-1][0], coords[j-1][1]), (coords[j][0], coords[j][1])).miles
                    except Exception as e:
                        logging.warning(f"Geodesic failed at point {j}: {str(e)}, using Haversine")
                        dist_miles = haversine_distance(coords[j-1][0], coords[j-1][1], coords[j][0], coords[j][1])
                    current_mp += dist_miles
                    file_mileposts.append(current_mp)
                    file_elevations.append(coords[j][2])
                mileposts.extend(file_mileposts)
                elevations.extend(file_elevations)
                segment_bounds.append({"name": os.path.basename(file_path), "start_mp": seg_start_mp, "end_mp": current_mp, "invert": invert_segment})
            except Exception as e:
                logging.error(f"Failed to parse TXT {file_path}: {str(e)}")
                raise ValueError(f"Failed to parse TXT {file_path}: {str(e)}")
        
        else:
            logging.warning(f"Unsupported file type: {file_path}")
            raise ValueError(f"Unsupported file type: {file_path}")
    
    if not mileposts:
        raise ValueError("No valid data extracted from files.")
    
    elevations = np.array(elevations)
    if elevation_units == "Meters":
        elevations *= 3.28084
    
    mileposts = np.array(mileposts)
    sorted_idx = np.argsort(mileposts)
    mileposts = mileposts[sorted_idx]
    elevations = elevations[sorted_idx]
    unique_mask = np.diff(mileposts, prepend=mileposts[0] - 1) > 0
    mileposts = mileposts[unique_mask]
    elevations = elevations[unique_mask]
    
    profile_start_mp = mileposts.min()
    profile_end_mp = mileposts.max()
    
    purge_mileposts = mileposts.copy()
    system_mileposts = mileposts.copy()
    system_elevations = elevations.copy()
    
    logging.info(f"Profile processed: Start MP={profile_start_mp:.1f}, End MP={profile_end_mp:.1f}, Points={len(mileposts)}")
    action = preview_profile_dialog(dialog_root, system_mileposts, system_elevations, segment_bounds=segment_bounds, elevation_units_label="ft")
    if action == "approve":
        return purge_mileposts, elevations, system_mileposts, system_elevations, profile_start_mp, profile_end_mp
    elif action == "reselect":
        return process_elevation_profile(dialog_root, initial_inputs)
    else:
        raise ValueError("Profile preview cancelled.")

def main():
    logging.info("Starting Pipeline Pigging Simulation")
    print("Starting Pipeline Pigging Simulation")
    try:
        dialog_root = tk.Tk()
        # Keep a mapped (but invisible) root so subsequent Toplevel dialogs reliably appear on Windows.
        dialog_root.title('Pipeline Pigging Simulation')
        dialog_root.geometry('1x1+0+0')
        try:
            dialog_root.attributes('-alpha', 0.0)
        except Exception:
            # Fallback for environments that don't support alpha
            dialog_root.withdraw()
        dialog_root.update_idletasks()
        initial_inputs = {
            'elevation_format': 'TXT',
            'elevation_units': 'Feet',
            'purge_start_mp': 0
        }
        logging.info("Calling process_elevation_profile")
        purge_mileposts, elevations, system_mileposts, system_elevations, profile_start_mp, profile_end_mp = process_elevation_profile(dialog_root, initial_inputs)
        logging.info("Calling get_user_inputs")
        inputs = get_user_inputs(dialog_root, profile_start_mp, profile_end_mp)
        inputs['elevation_format'] = initial_inputs['elevation_format']
        inputs['elevation_units'] = initial_inputs['elevation_units']
        
        print_inputs(inputs)
        
        if len(system_mileposts) > inputs['resolution'] * 2:
            new_system_mp = np.linspace(profile_start_mp, profile_end_mp, inputs['resolution'] * 2)
            cs_system = CubicSpline(system_mileposts, system_elevations)
            system_elevations = cs_system(new_system_mp)
            system_mileposts = new_system_mp
        
        mask = (purge_mileposts >= inputs['purge_start_mp']) & (purge_mileposts <= inputs['purge_end_mp'])
        purge_mileposts = purge_mileposts[mask]
        elevations = elevations[mask]



        # Purge segment resolution handling
        # - If resample_profile_to_resolution is True (default): resample to exactly `resolution` points using a spline.
        # - If False: use native uploaded points within Purge Start/End, but ensure start/end are included.
        resample_to_resolution = bool(inputs.get('resample_profile_to_resolution', True))
        cs_purge = CubicSpline(system_mileposts, system_elevations, extrapolate=True)

        if resample_to_resolution:
            try:
                purge_res = int(inputs.get('resolution', 500))
            except Exception:
                purge_res = 500
            purge_res = max(2, purge_res)

            new_purge_mp = np.linspace(inputs['purge_start_mp'], inputs['purge_end_mp'], purge_res)
            elevations = cs_purge(new_purge_mp)
            purge_mileposts = new_purge_mp

            logging.info(
                f"Resampled purge segment to {purge_res} points "
                f"from MP {inputs['purge_start_mp']:.3f} to {inputs['purge_end_mp']:.3f}"
            )
        else:
            # Use native points, but ensure boundaries exist (and avoid edge-case empty masks).
            if purge_mileposts.size == 0:
                purge_mileposts = np.array([inputs['purge_start_mp'], inputs['purge_end_mp']])
                elevations = cs_purge(purge_mileposts)
            else:
                tol = 1e-9
                if purge_mileposts[0] - inputs['purge_start_mp'] > tol:
                    purge_mileposts = np.insert(purge_mileposts, 0, inputs['purge_start_mp'])
                    elevations = np.insert(elevations, 0, cs_purge(inputs['purge_start_mp']))
                if inputs['purge_end_mp'] - purge_mileposts[-1] > tol:
                    purge_mileposts = np.append(purge_mileposts, inputs['purge_end_mp'])
                    elevations = np.append(elevations, cs_purge(inputs['purge_end_mp']))

            logging.info(
                f"Using native purge profile points: {len(purge_mileposts)} points "
                f"from MP {purge_mileposts[0]:.3f} to {purge_mileposts[-1]:.3f}"
            )

        while True:

            logging.info("Running simulation")
            # Optional N2 cutoff default (None = no cutoff)
            inputs.setdefault('n2_optional_cutoff_scf', None)

            results = run_simulation(dialog_root, inputs, purge_mileposts, elevations, system_mileposts, system_elevations)
            visualize_results(results, inputs)
            print_condensed_table(results, inputs)
            
            print_alarm_summary(results, inputs)
            resp = messagebox.askyesnocancel("Output Satisfactory", "Is the output satisfactory?", parent=dialog_root)
            if resp is True:
                export_results(dialog_root, inputs, results)
                break
            elif resp is None:
                logging.info("User cancelled after reviewing outputs; exiting without export.")
                break
            else:
                try:
                    dialog_root.update()
                    # --- Revision prompt: TOTAL N2 injected (volume cutoff) ---
                    last_total_n2 = None
                    try:
                        last_i = int(results.get('last_valid_i', -1))
                        if ('cumulative_n2' in results) and (last_i is not None) and (last_i >= 0) and (last_i < len(results['cumulative_n2'])):
                            last_total_n2 = float(results['cumulative_n2'][last_i])
                    except Exception:
                        last_total_n2 = None
                    cutoff_current = inputs.get('n2_cutoff_scf', None)
                    cutoff_disp = 'auto' if (cutoff_current is None) else f"{float(cutoff_current):.0f}"
                    if last_total_n2 is None:
                        prompt = f"New TOTAL N2 injected cutoff (SCF) [blank=keep, 'auto'=clear]\nLast setting: {cutoff_disp}"
                    else:
                        prompt = f"New TOTAL N2 injected cutoff (SCF) [blank=keep, 'auto'=clear]\nLast setting: {cutoff_disp}\nLast run injected: {last_total_n2:.0f} SCF"
                    new_cutoff_str = tk.simpledialog.askstring('Adjust N2 Injection', prompt, parent=dialog_root)
                    if new_cutoff_str is not None:
                        s = str(new_cutoff_str).strip().lower()
                        if s == '':
                            pass
                        elif s in ('auto', 'none', 'clear', 'off'):
                            inputs['n2_cutoff_scf'] = None
                        else:
                            inputs['n2_cutoff_scf'] = float(s)

                    new_n2_rate = tk.simpledialog.askfloat(
                        "Input", f"New Max N2 Rate (SCFM) [optional] (Last={inputs.get('max_n2_rate_scfm', None)}):",
                        parent=dialog_root, minvalue=0.1
                    )
                    if new_n2_rate is not None:
                        inputs['max_n2_rate_scfm'] = new_n2_rate
                    new_exit_run = tk.simpledialog.askfloat(
                        "Input", f"New Min Exit Pressure (Run) (Last={inputs['exit_pressure_run']:.1f}):",
                        parent=dialog_root, minvalue=0
                    )
                    if new_exit_run is not None:
                        inputs['exit_pressure_run'] = new_exit_run
                    new_exit_end = tk.simpledialog.askfloat(
                        "Input", f"New Exit Pressure (End) (Last={inputs['exit_pressure_end']:.1f}):",
                        parent=dialog_root, minvalue=0
                    )
                    if new_exit_end is not None:
                        inputs['exit_pressure_end'] = new_exit_end
                    new_target_speed = tk.simpledialog.askfloat(
                        "Input", f"New Target Pig Speed (mph) (Last={inputs['target_pig_speed']:.1f}):",
                        parent=dialog_root, minvalue=0.1, maxvalue=inputs['max_pig_speed']
                    )
                    if new_target_speed is not None:
                        inputs['target_pig_speed'] = new_target_speed
                    dialog_root.update()
                    print_inputs(inputs)
                    continue
                except ValueError as e:
                    logging.error(f"Input revision error: {str(e)}")
                    messagebox.showerror("Input Error", f"Invalid input: {str(e)}", parent=dialog_root)
    
    except Exception as e:
        error_msg = f"Simulation failed: {str(e)}"
        logging.error(error_msg)
        print(error_msg)
        try:
            messagebox.showerror("Error", error_msg, parent=dialog_root)
        except Exception as tk_e:
            logging.error(f"Failed to show error messagebox: {str(tk_e)}")
            print(f"Error: Failed to show error messagebox: {str(tk_e)}")
        raise
    finally:
        dialog_root.destroy()

class TestPurgeSimulation(unittest.TestCase):
    def test_friction_loss(self):
        v, L, D, rho, mu, eps = 10, 1000, 0.5, 62.4, 1e-5, 0.00015
        loss = calculate_friction_loss(v, L, D, rho, mu, eps)
        self.assertGreater(loss, 0, "Friction loss should be positive.")
    
    def test_trendline_slope(self):
        mileposts = np.array([0, 1, 2])
        elevations = np.array([100, 110, 120])
        slope = calculate_trendline_slope(mileposts, elevations, 0, 2)
        self.assertEqual(slope, 10, "Slope should be 10 ft/mile.")

if __name__ == "__main__":
    main()